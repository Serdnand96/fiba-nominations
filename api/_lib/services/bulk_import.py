import io
import csv
import re
from openpyxl import load_workbook
from api._lib.database import supabase

COLUMN_MAP = {
    "nombre": "name",
    "name": "name",
    "email": "email",
    "país": "country",
    "pais": "country",
    "country": "country",
    "teléfono": "phone",
    "telefono": "phone",
    "phone": "phone",
    "pasaporte": "passport",
    "passport": "passport",
    "rol": "role",
    "role": "role",
}

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def _read_csv(file_bytes: bytes) -> list[dict]:
    """Parse CSV bytes into a list of row dicts with normalized column names."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = {_norm(h): COLUMN_MAP.get(_norm(h), _norm(h)) for h in (reader.fieldnames or [])}
    rows = []
    for row in reader:
        mapped = {}
        for orig_key, val in row.items():
            mapped_key = headers.get(_norm(orig_key), _norm(orig_key))
            mapped[mapped_key] = val
        rows.append(mapped)
    return rows


def _read_xlsx(file_bytes: bytes) -> list[dict]:
    """Parse XLSX bytes into a list of row dicts with normalized column names."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        return []
    headers = [COLUMN_MAP.get(_norm(str(h or "")), _norm(str(h or ""))) for h in raw_headers]
    rows = []
    for row_vals in rows_iter:
        row = {}
        for i, val in enumerate(row_vals):
            if i < len(headers):
                row[headers[i]] = str(val) if val is not None else ""
        rows.append(row)
    return rows


def _norm(s: str) -> str:
    return s.strip().lower()


def _clean(val: str | None) -> str | None:
    if val is None:
        return None
    s = val.strip()
    return s if s and s.lower() not in ("none", "nan", "") else None


def process_bulk_import(file_bytes: bytes, filename: str) -> dict:
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "csv":
        rows = _read_csv(file_bytes)
    elif ext in ("xlsx", "xls"):
        rows = _read_xlsx(file_bytes)
    else:
        return {
            "total": 0, "imported": 0, "skipped": 0,
            "errors": [{"row": 0, "email": "", "reason": f"Unsupported file type: .{ext}"}],
        }

    if not rows:
        return {"total": 0, "imported": 0, "skipped": 0, "errors": []}

    # Check required columns exist
    sample_keys = set(rows[0].keys())
    for req in ["name", "email", "role"]:
        if req not in sample_keys:
            return {
                "total": len(rows), "imported": 0, "skipped": 0,
                "errors": [{"row": 0, "email": "", "reason": f"Missing required column: {req}"}],
            }

    # Get existing emails from DB
    existing = supabase.table("personnel").select("email").execute()
    existing_emails = {r["email"].lower() for r in existing.data}

    errors = []
    valid_rows = []
    skipped = 0

    for idx, row in enumerate(rows):
        row_num = idx + 2  # 1-indexed + header row
        name = (row.get("name") or "").strip()
        email = (row.get("email") or "").strip()
        role = (row.get("role") or "").strip().upper()

        if not name or name.lower() == "nan":
            errors.append({"row": row_num, "email": email, "reason": "Name is required"})
            continue
        if not email or email.lower() == "nan":
            errors.append({"row": row_num, "email": email, "reason": "Email is required"})
            continue
        if not EMAIL_REGEX.match(email):
            errors.append({"row": row_num, "email": email, "reason": "Invalid email format"})
            continue
        if role not in ("VGO", "TD"):
            errors.append({"row": row_num, "email": email, "reason": f"Role must be VGO or TD, got '{role}'"})
            continue
        if email.lower() in existing_emails:
            skipped += 1
            continue

        existing_emails.add(email.lower())
        record = {
            "name": name,
            "email": email,
            "role": role,
            "country": _clean(row.get("country")),
            "phone": _clean(row.get("phone")),
            "passport": _clean(row.get("passport")),
        }
        valid_rows.append(record)

    imported = 0
    if valid_rows:
        result = supabase.table("personnel").insert(valid_rows).execute()
        imported = len(result.data)

    return {
        "total": len(rows),
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
    }
