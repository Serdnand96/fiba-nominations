"""Import y export de las planillas de logística de FIBA Americas.

Dos formatos, los dos escritos a mano por la organización y los dos
irregulares:

* **Flight Manifest** (`hoja Staff`) — llegadas a la izquierda y salidas a la
  derecha de la misma fila, headers en la fila 5. Además una hoja `Teams` con
  delegaciones (una fila = un grupo con headcount, no una persona).
* **Rooming List** — headers en la fila 2, una columna por noche con `1`/`OUT`,
  y las celdas `Room` combinadas verticalmente cuando dos personas comparten
  habitación.

El parseo es deliberadamente tolerante: fechas `30.07.26` conviven con
`datetime` reales, las horas aparecen como `08:50 am`, `22:30 PM`, `07:00:00` y
hasta `12:OO` (con letra O), y hay filas donde `Names` y `Last Name` están
invertidas. Lo que no se puede interpretar se reporta como warning en el
preview en vez de romper el import o, peor, de inventar un valor.
"""
import io
import re
import unicodedata
from datetime import date, datetime, time
from difflib import SequenceMatcher
from typing import Optional

from openpyxl import Workbook, load_workbook

from api._lib.database import supabase

# ── Normalización y parseo de celdas ─────────────────────────────────────────

_PLACEHOLDERS = {"", "n/a", "na", "tbc", "tbd", "none", "nan", "-", "?"}


def norm(value) -> str:
    """Texto comparable: sin acentos, sin dobles espacios, en minúsculas."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text)


def clean(value) -> Optional[str]:
    """Valor de celda listo para guardar, o None si es vacío/placeholder."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if norm(text) in _PLACEHOLDERS:
        return None
    return re.sub(r"\s+", " ", text) or None


def parse_date(value, default_year: Optional[int] = None) -> Optional[str]:
    """Fecha en ISO, o None. Acepta datetime, `30.07.26`, `30/07/2026`, `30-07`."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if norm(text) in _PLACEHOLDERS:
        return None

    match = re.match(r"^(\d{1,2})[./-](\d{1,2})(?:[./-](\d{2,4}))?$", text)
    if not match:
        return None
    day, month, year = match.group(1), match.group(2), match.group(3)
    if year is None:
        if default_year is None:
            return None
        year_num = default_year
    else:
        year_num = int(year)
        if year_num < 100:  # "26" → 2026
            year_num += 2000
    try:
        return date(year_num, int(month), int(day)).isoformat()
    except ValueError:
        return None


def parse_time(value) -> Optional[str]:
    """Hora en HH:MM, o None.

    Tolera `08:50 am`, `22:30 PM` (sí, 22 con PM: se ignora el sufijo cuando la
    hora ya es de 24), `07:00:00`, `1530` y el típico `12:OO` tipeado con letra
    O en lugar de cero.
    """
    if value is None:
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")

    text = str(value).strip()
    if norm(text) in _PLACEHOLDERS:
        return None

    lowered = text.lower()
    is_pm = "pm" in lowered
    is_am = "am" in lowered
    # Los ceros tipeados como O sólo se corrigen dentro de lo que parece hora.
    body = re.sub(r"[ap]\.?m\.?", "", lowered).strip()
    body = body.replace("o", "0").replace(".", ":")

    match = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", body)
    if match:
        hour, minute = int(match.group(1)), int(match.group(2))
    else:
        digits = re.match(r"^(\d{1,2})(\d{2})$", body)
        if digits:
            hour, minute = int(digits.group(1)), int(digits.group(2))
        else:
            return None

    if is_pm and hour < 12:
        hour += 12
    if is_am and hour == 12:
        hour = 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return f"{hour:02d}:{minute:02d}"


def guess_transport_mode(airport: Optional[str], flight: Optional[str]) -> str:
    """El manifest mete micros y ferries en las columnas de vuelo.

    Los ferries se evalúan primero a propósito: "Buquebus" contiene la
    subcadena "bus" y si no, todo el que cruza el Río de la Plata viaja en
    micro.
    """
    blob = norm(f"{airport or ''} {flight or ''}")
    if any(k in blob for k in ("buquebus", "colonia express", "puerto", "ferry")):
        return "ferry"
    if "bus" in blob or "terminal" in blob:
        return "bus"
    return "flight"


# ── Matching contra personnel / employees ────────────────────────────────────

_MATCH_THRESHOLD = 0.87   # abajo de esto no se vincula solo
_REVIEW_THRESHOLD = 0.70  # entre medio, se muestra como dudoso


_TOKEN_THRESHOLD = 0.80  # dos palabras son "la misma" arriba de esto


def _name_key(text: str) -> frozenset:
    return frozenset(w for w in norm(text).split() if len(w) > 1)


def _shared_words(words_a: frozenset, words_b: frozenset) -> int:
    """Palabras en común, tolerando un typo por palabra.

    "BUELVAS" y "Vuelvas" son la misma persona escrita por dos personas
    distintas. Pero un typo en una palabra corta es indistinguible de un
    apellido diferente ("Lima"/"Lina" puntúan igual que "Juyo"/"Guyo"), así que
    el parecido difuso solo cuenta si además hay al menos una palabra idéntica
    — el nombre de pila haciendo de ancla.
    """
    exact = words_a & words_b
    if not exact:
        return 0

    shared = len(exact)
    rest_a = [w for w in words_a - exact]
    rest_b = [w for w in words_b - exact]
    used = set()
    for wa in rest_a:
        for wb in rest_b:
            if wb in used:
                continue
            if SequenceMatcher(None, wa, wb).ratio() >= _TOKEN_THRESHOLD:
                shared += 1
                used.add(wb)
                break
    return shared


def _possible_duplicate(a: str, b: str) -> bool:
    """¿Podrían ser la misma persona escrita distinto, sin llegar a fusionarse?

    No sirve reusar `_similarity` para esto: sobre nombres cortos el parecido
    carácter a carácter es ruidoso — "Gabriela Schaer" y "Gabriel Guimaraes"
    puntúan más alto que "Jose Juyo" y "JOSE GUYO", que sí son la misma
    persona. La señal buena es más específica: comparten al menos una palabra
    exacta (el nombre de pila haciendo de ancla) y lo que queda se parece lo
    suficiente como para ser un typo.
    """
    words_a, words_b = _name_key(a), _name_key(b)
    exact = words_a & words_b
    if not exact:
        return False
    rest_a, rest_b = words_a - exact, words_b - exact
    if not rest_a or not rest_b:
        return False
    return any(
        SequenceMatcher(None, wa, wb).ratio() >= 0.6
        for wa in rest_a for wb in rest_b
    )


def _similarity(a: str, b: str) -> float:
    """Parecido entre dos nombres, sin importar el orden de las palabras.

    Tres cosas pasan seguido en estas planillas y las tres tienen que dar alto:

    * el orden se invierte — "SCHAER GABRIELA" vs "Gabriela Schaer";
    * la planilla acorta el nombre legal — "Gabriela Schaer" es la misma
      persona que "Gabriela Schaer Araya", y por eso el solapamiento se divide
      por el nombre *más corto* y no por el más largo;
    * hay typos — "Guyo"/"Juyo", "Varella"/"Varela", que los levanta el
      SequenceMatcher sobre las palabras ordenadas.

    La versión por contención pide al menos dos palabras en el nombre corto:
    con una sola, "Daniel" entraría en cualquier Daniel del padrón.
    """
    words_a, words_b = _name_key(a), _name_key(b)
    if not words_a or not words_b:
        return 0.0

    shared = _shared_words(words_a, words_b)
    shorter = min(len(words_a), len(words_b))
    containment = shared / shorter if shorter >= 2 else 0.0

    sorted_a = " ".join(sorted(words_a))
    sorted_b = " ".join(sorted(words_b))
    fuzzy = SequenceMatcher(None, sorted_a, sorted_b).ratio()
    return max(containment, fuzzy)


class Directory:
    """Padrón de referencia: personnel + employees, para vincular filas."""

    def __init__(self):
        self.personnel = (
            supabase.table("personnel")
            .select("id,name,country,passport,phone,email,role")
            .execute()
            .data
        ) or []
        self.employees = (
            supabase.table("employees")
            .select("id,name,email,position,phone")
            .execute()
            .data
        ) or []

    def match(self, name: str) -> dict:
        """→ {kind, id, matched_name, score, status, rival}.

        status: 'matched' (se vincula), 'review' (parecido pero dudoso, se
        importa sin vincular y se avisa) o 'new' (nadie se parece).

        Un empate arriba baja a 'review' aunque el puntaje alcance: si dos
        personas del padrón puntúan casi igual, elegir una es adivinar, y en
        una planilla de vuelos elegir mal es mandar a alguien al aeropuerto
        equivocado.
        """
        scored = []
        for kind, rows in (("personnel", self.personnel), ("employee", self.employees)):
            for row in rows:
                score = _similarity(name, row.get("name") or "")
                if score > 0:
                    scored.append({
                        "kind": kind,
                        "id": row["id"],
                        "matched_name": row.get("name"),
                        "score": round(score, 3),
                    })
        if not scored:
            return {"kind": None, "id": None, "matched_name": None,
                    "score": 0.0, "status": "new", "rival": None}

        scored.sort(key=lambda c: c["score"], reverse=True)
        best = scored[0]
        rival = next((c for c in scored[1:] if best["score"] - c["score"] <= 0.02), None)

        if rival is not None and best["score"] >= _REVIEW_THRESHOLD:
            return {**best, "status": "review", "rival": rival["matched_name"]}
        if best["score"] >= _MATCH_THRESHOLD:
            return {**best, "status": "matched", "rival": None}
        if best["score"] >= _REVIEW_THRESHOLD:
            return {**best, "status": "review", "rival": None}
        return {"kind": None, "id": None, "matched_name": None,
                "score": best["score"], "status": "new", "rival": None}


def review_warning(row_idx: int, name: str, match: dict) -> dict:
    """Aviso de match dudoso, para que lo resuelva una persona."""
    if match.get("rival"):
        message = (f"'{name}' puntúa casi igual contra '{match['matched_name']}' y "
                   f"'{match['rival']}' — se importa sin vincular")
    else:
        message = (f"'{name}' se parece a '{match['matched_name']}' "
                   f"({int(match['score'] * 100)}%) — se importa sin vincular")
    return {"row": row_idx, "field": "name", "message": message}


def categorize(role: Optional[str]) -> str:
    """Categoría del padrón a partir del cargo escrito en la planilla."""
    blob = norm(role)
    if not blob:
        return "official"
    if "vip" in blob or "autoridad" in blob or "presidente" in blob or "pte" in blob:
        return "vip"
    if blob.startswith("fiba") or "comms" in blob or "comunicaciones" in blob \
            or "manager" in blob or "support" in blob or "staff" in blob:
        return "fiba_staff"
    return "official"


# ── Lectura del Flight Manifest ──────────────────────────────────────────────

def _find_header_row(ws, needles: tuple[str, ...], limit: int = 12) -> Optional[int]:
    """Fila de headers: la primera que contiene todas las palabras buscadas.

    Las planillas arrancan con un título y filas en blanco, así que asumir la
    fila 1 no sirve.
    """
    for row_idx in range(1, min(ws.max_row, limit) + 1):
        values = {norm(c.value) for c in ws[row_idx] if c.value is not None}
        if all(any(needle in v for v in values) for needle in needles):
            return row_idx
    return None


def _header_columns(ws, header_row: int) -> dict:
    """{texto normalizado del header: índice de columna}."""
    out = {}
    for cell in ws[header_row]:
        key = norm(cell.value)
        if key and key not in out:
            out[key] = cell.column
    return out


def _pick(columns: dict, *needles: str) -> Optional[int]:
    """Columna cuyo header contiene alguno de los textos dados."""
    for needle in needles:
        for header, col in columns.items():
            if needle in header:
                return col
    return None


def _cell(ws, row: int, col: Optional[int]):
    return ws.cell(row=row, column=col).value if col else None


def _competition_year(competition_id: str) -> Optional[int]:
    rows = (
        supabase.table("competitions")
        .select("start_date")
        .eq("id", competition_id)
        .execute()
        .data
    )
    if rows and rows[0].get("start_date"):
        try:
            return date.fromisoformat(rows[0]["start_date"]).year
        except ValueError:
            return None
    return None


def read_manifest(content: bytes, competition_id: str) -> dict:
    """Parsea el manifest y devuelve filas + avisos, sin tocar la base."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    year = _competition_year(competition_id)
    directory = Directory()

    rows: list[dict] = []
    warnings: list[dict] = []

    staff_sheet = next(
        (wb[n] for n in wb.sheetnames if norm(n) in ("staff", "fiba staff")),
        wb.worksheets[0],
    )
    header_row = _find_header_row(staff_sheet, ("name", "role"))
    if header_row is None:
        return {
            "rows": [],
            "warnings": [{"row": 0, "field": "header",
                          "message": "No se encontró la fila de encabezados (se buscaron 'Name' y 'Role')"}],
            "sheet": staff_sheet.title,
        }

    columns = _header_columns(staff_sheet, header_row)

    # Llegadas (izquierda) y salidas (derecha) comparten la fila. Las dos mitades
    # repiten headers ("Name", "Flight #"), así que se parten por la columna del
    # header de salidas: _header_columns se queda con la primera aparición, y la
    # segunda se busca a mano.
    col_first_name = _pick(columns, "name")
    col_last_name = _pick(columns, "last name")
    col_role = _pick(columns, "role")
    col_passport = _pick(columns, "passport")
    col_arr_date = _pick(columns, "arrival date")
    col_arr_time = _pick(columns, "arrival time")
    col_arr_flight = _pick(columns, "flight #")
    col_arr_airport = _pick(columns, "arriving airport")
    col_arr_comments = _pick(columns, "comment")

    col_dep_date = _pick(columns, "departure date")
    col_dep_pickup = _pick(columns, "pick up time")
    col_dep_flight_time = _pick(columns, "departure flight")
    col_dep_airport = _pick(columns, "departing airport")
    # El "Flight #" de salida es el que está a la derecha del bloque de salidas.
    col_dep_flight = None
    if col_dep_date:
        for cell in staff_sheet[header_row]:
            if cell.column > col_dep_date and "flight #" in norm(cell.value):
                col_dep_flight = cell.column
                break
    col_dep_comments = None
    if col_dep_date:
        for cell in staff_sheet[header_row]:
            if cell.column > col_dep_date and "comment" in norm(cell.value):
                col_dep_comments = cell.column
                break

    for row_idx in range(header_row + 1, staff_sheet.max_row + 1):
        raw_first = clean(_cell(staff_sheet, row_idx, col_first_name))
        raw_last = clean(_cell(staff_sheet, row_idx, col_last_name))
        role = clean(_cell(staff_sheet, row_idx, col_role))
        if not raw_first and not raw_last:
            continue
        # "TBC / TBC" es un lugar reservado sin persona todavía; clean() lo
        # vuelve None y la fila queda descartada salvo que traiga cargo.
        if not raw_first and not raw_last and not role:
            continue

        # Mismas planillas que la rooming: Names / Last Name vienen invertidos en
        # varias filas. Decidir el orden por cuál combinación se parece más a
        # alguien del padrón, no por el header (antes read_manifest confiaba en
        # las columnas y guardaba nombre/apellido cambiados de lugar).
        direct = " ".join(x for x in (raw_first, raw_last) if x)
        swapped = " ".join(x for x in (raw_last, raw_first) if x)
        match_direct = directory.match(direct)
        match_swapped = directory.match(swapped)
        if match_swapped["score"] > match_direct["score"]:
            first, last, match = raw_last, raw_first, match_swapped
        else:
            first, last, match = raw_first, raw_last, match_direct
        name = " ".join(x for x in (first, last) if x)
        if match["status"] == "review":
            warnings.append(review_warning(row_idx, name, match))

        arr_date_raw = _cell(staff_sheet, row_idx, col_arr_date)
        arr_date = parse_date(arr_date_raw, year)
        if arr_date_raw is not None and arr_date is None and clean(arr_date_raw):
            warnings.append({"row": row_idx, "field": "arrival_date",
                             "message": f"Fecha de llegada ilegible: '{arr_date_raw}'"})

        arr_time_raw = _cell(staff_sheet, row_idx, col_arr_time)
        arr_time = parse_time(arr_time_raw)
        if arr_time_raw is not None and arr_time is None and clean(arr_time_raw):
            warnings.append({"row": row_idx, "field": "arrival_time",
                             "message": f"Hora de llegada ilegible: '{arr_time_raw}'"})

        dep_date_raw = _cell(staff_sheet, row_idx, col_dep_date)
        dep_date = parse_date(dep_date_raw, year)
        if dep_date_raw is not None and dep_date is None and clean(dep_date_raw):
            warnings.append({"row": row_idx, "field": "departure_date",
                             "message": f"Fecha de salida ilegible: '{dep_date_raw}'"})

        dep_time_raw = _cell(staff_sheet, row_idx, col_dep_flight_time)
        dep_time = parse_time(dep_time_raw)
        if dep_time_raw is not None and dep_time is None and clean(dep_time_raw):
            warnings.append({"row": row_idx, "field": "departure_time",
                             "message": f"Hora de salida ilegible: '{dep_time_raw}'"})

        pickup_raw = _cell(staff_sheet, row_idx, col_dep_pickup)
        pickup = parse_time(pickup_raw)
        if pickup_raw is not None and pickup is None and clean(pickup_raw):
            warnings.append({"row": row_idx, "field": "pickup_time",
                             "message": f"Hora de pick-up ilegible: '{pickup_raw}'"})

        arr_airport = clean(_cell(staff_sheet, row_idx, col_arr_airport))
        arr_flight = clean(_cell(staff_sheet, row_idx, col_arr_flight))
        dep_airport = clean(_cell(staff_sheet, row_idx, col_dep_airport))
        dep_flight = clean(_cell(staff_sheet, row_idx, col_dep_flight))

        legs = []
        if any([arr_date, arr_time, arr_flight, arr_airport]):
            legs.append({
                "direction": "arrival",
                "date": arr_date,
                "time": arr_time,
                "pickup_time": None,
                "flight_number": arr_flight,
                "carrier": None,
                "airport": arr_airport,
                "transport_mode": guess_transport_mode(arr_airport, arr_flight),
                "comments": clean(_cell(staff_sheet, row_idx, col_arr_comments)),
            })
        if any([dep_date, dep_time, dep_flight, dep_airport, pickup]):
            legs.append({
                "direction": "departure",
                "date": dep_date,
                "time": dep_time,
                "pickup_time": pickup,
                "flight_number": dep_flight,
                "carrier": None,
                "airport": dep_airport,
                "transport_mode": guess_transport_mode(dep_airport, dep_flight),
                "comments": clean(_cell(staff_sheet, row_idx, col_dep_comments)),
            })

        passport = clean(_cell(staff_sheet, row_idx, col_passport))

        # Debajo de la tabla la planilla sigue con etiquetas de sección ("VIP")
        # y un pie de contactos ("Ricardo Serra: +1 …"), que en la columna de
        # nombre son indistinguibles de una persona. Un participante de verdad
        # trae además cargo, pasaporte o algún vuelo; un rótulo no trae nada.
        if not role and not passport and not legs:
            continue

        rows.append({
            "source_row": row_idx,
            "first_name": first or last or "?",
            "last_name": last if first else None,
            "role": role,
            "category": categorize(role),
            "passport": passport,
            "pax": 1,
            "match": match,
            "legs": legs,
        })

    # Hoja Teams: delegaciones con headcount, no personas.
    teams_sheet = next((wb[n] for n in wb.sheetnames if norm(n) == "teams"), None)
    if teams_sheet is not None:
        team_header = _find_header_row(teams_sheet, ("teams",))
        if team_header:
            tcols = _header_columns(teams_sheet, team_header)
            col_team = _pick(tcols, "teams")
            col_tdate = _pick(tcols, "departure date")
            col_tflight = _pick(tcols, "departure flight")
            col_tnumber = _pick(tcols, "flight #")
            col_tpax = _pick(tcols, "pax")
            col_tairport = _pick(tcols, "departing airport")
            col_thotel = _pick(tcols, "departure hotel")

            for row_idx in range(team_header + 1, teams_sheet.max_row + 1):
                team = clean(_cell(teams_sheet, row_idx, col_team))
                if not team:
                    continue
                pax_raw = _cell(teams_sheet, row_idx, col_tpax)
                try:
                    pax = max(int(float(pax_raw)), 1) if pax_raw not in (None, "") else 1
                except (TypeError, ValueError):
                    pax = 1
                airport = clean(_cell(teams_sheet, row_idx, col_tairport))
                flight = clean(_cell(teams_sheet, row_idx, col_tnumber))
                rows.append({
                    "source_row": row_idx,
                    "sheet": teams_sheet.title,
                    "first_name": team,
                    "last_name": None,
                    "role": "Delegación",
                    "category": "team",
                    "passport": None,
                    "pax": pax,
                    "match": {"kind": None, "id": None, "matched_name": None,
                              "score": 0.0, "status": "new"},
                    "legs": [{
                        "direction": "departure",
                        "date": parse_date(_cell(teams_sheet, row_idx, col_tdate), year),
                        "time": parse_time(_cell(teams_sheet, row_idx, col_tflight)),
                        "pickup_time": parse_time(_cell(teams_sheet, row_idx, col_thotel)),
                        "flight_number": flight,
                        "carrier": None,
                        "airport": airport,
                        "transport_mode": guess_transport_mode(airport, flight),
                        "comments": None,
                    }],
                })

    return {"rows": rows, "warnings": warnings, "sheet": staff_sheet.title}


# ── Lectura de la Rooming List ───────────────────────────────────────────────

def _room_cell(ws, col: int, row: int) -> tuple[Optional[str], Optional[int]]:
    """(tipo de habitación, fila que la encabeza si es compartida).

    Una "Double" cuya celda está combinada verticalmente es la forma que tiene
    la planilla de decir que esas filas comparten habitación. La fila donde
    arranca el bloque es la titular; las de abajo son sus compañeras.
    """
    for merged in ws.merged_cells.ranges:
        if merged.min_col <= col <= merged.max_col and merged.min_row <= row <= merged.max_row:
            value = clean(ws.cell(row=merged.min_row, column=merged.min_col).value)
            if merged.max_row == merged.min_row:
                return value, None
            return value, (None if row == merged.min_row else merged.min_row)
    return clean(ws.cell(row=row, column=col).value), None


def read_rooming(content: bytes, competition_id: str) -> dict:
    """Parsea la rooming list y devuelve filas + avisos, sin tocar la base."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    year = _competition_year(competition_id)
    directory = Directory()

    ws = next(
        (wb[n] for n in wb.sheetnames if "staff" in norm(n) or "rooming" in norm(n)),
        wb.worksheets[0],
    )
    header_row = _find_header_row(ws, ("name", "room"))
    if header_row is None:
        return {
            "rows": [],
            "warnings": [{"row": 0, "field": "header",
                          "message": "No se encontró la fila de encabezados (se buscaron 'Name' y 'Room')"}],
            "sheet": ws.title,
        }

    columns = _header_columns(ws, header_row)
    col_a = _pick(columns, "name")
    col_b = _pick(columns, "last name")
    col_role = _pick(columns, "role")
    col_room = _pick(columns, "room")
    col_diet = _pick(columns, "dietary", "restric")
    col_comments = _pick(columns, "comentario", "comment")

    # Columnas de noche: las que tienen una fecha por header.
    night_columns: list[tuple[int, str]] = []
    for cell in ws[header_row]:
        parsed = parse_date(cell.value, year)
        if parsed:
            night_columns.append((cell.column, parsed))
    night_columns.sort(key=lambda c: c[0])

    rows: list[dict] = []
    warnings: list[dict] = []
    if not night_columns:
        warnings.append({"row": header_row, "field": "nights",
                         "message": "No se encontraron columnas de noche con fecha en el encabezado"})

    by_source_row: dict = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_a = clean(_cell(ws, row_idx, col_a))
        raw_b = clean(_cell(ws, row_idx, col_b))
        role = clean(_cell(ws, row_idx, col_role))
        if not raw_a and not raw_b:
            continue

        # En esta planilla las columnas Names / Last Name están invertidas en
        # varias filas, así que el orden se decide por cuál combinación se
        # parece más a alguien del padrón, no por el header.
        direct = " ".join(x for x in (raw_a, raw_b) if x)
        swapped = " ".join(x for x in (raw_b, raw_a) if x)
        match_direct = directory.match(direct)
        match_swapped = directory.match(swapped)
        if match_swapped["score"] > match_direct["score"]:
            first, last, match = raw_b, raw_a, match_swapped
        else:
            first, last, match = raw_a, raw_b, match_direct
        name = " ".join(x for x in (first, last) if x)

        if match["status"] == "review":
            warnings.append(review_warning(row_idx, name, match))

        # Noches: la primera columna con contenido es el check-in, la marcada
        # OUT (o el día siguiente a la última noche) es el check-out.
        check_in = None
        check_out = None
        last_night = None
        for col, day in night_columns:
            value = ws.cell(row=row_idx, column=col).value
            text = norm(value)
            if not text:
                continue
            if text == "out":
                check_out = day
            else:
                if check_in is None:
                    check_in = day
                last_night = day
        if check_out is None and last_night:
            check_out = date.fromordinal(date.fromisoformat(last_night).toordinal() + 1).isoformat()

        room_type, share_with_row = (None, None)
        if col_room:
            room_type, share_with_row = _room_cell(ws, col_room, row_idx)

        # La planilla carga las noches una sola vez por habitación, en la fila
        # titular; la compañera queda en blanco. Para el hotel eso es correcto
        # (se factura la habitación), pero la persona sí duerme esas noches, así
        # que hereda las fechas. El total sigue contando la habitación una vez:
        # lo hace `is_room_holder`.
        holder = by_source_row.get(share_with_row) if share_with_row else None
        if holder and not check_in:
            check_in = holder["stay"]["check_in"]
            check_out = check_out or holder["stay"]["check_out"]
            if not room_type:
                room_type = holder["stay"]["room_type"]

        record = {
            "source_row": row_idx,
            "first_name": first or last or "?",
            "last_name": last if first else None,
            "role": role,
            "category": categorize(role),
            "dietary_restrictions": clean(_cell(ws, row_idx, col_diet)),
            "match": match,
            "is_room_holder": share_with_row is None,
            "stay": {
                "room_type": room_type,
                "room_number": None,
                "check_in": check_in,
                "check_out": check_out,
                "comments": clean(_cell(ws, row_idx, col_comments)),
                # Fila de la planilla con la que comparte; se traduce a
                # share_with_id recién al confirmar, cuando existen los ids.
                "share_with_row": share_with_row,
            },
            "nights": (
                (date.fromisoformat(check_out) - date.fromisoformat(check_in)).days
                if check_in and check_out else 0
            ),
        }
        rows.append(record)
        by_source_row[row_idx] = record

    return {
        "rows": rows,
        "warnings": warnings,
        "sheet": ws.title,
        "dates": [d for _, d in night_columns],
        # Noches de habitación, que es lo que se le factura al hotel: la
        # compañera de cuarto no suma una segunda vez.
        "total_nights": sum(r["nights"] for r in rows if r["is_room_holder"]),
    }


# ── Preview (no escribe) ─────────────────────────────────────────────────────

def _summarize(rows: list[dict], warnings: list[dict], **extra) -> dict:
    return {
        "total": len(rows),
        "matched": sum(1 for r in rows if r["match"]["status"] == "matched"),
        "review": sum(1 for r in rows if r["match"]["status"] == "review"),
        "new": sum(1 for r in rows if r["match"]["status"] == "new"),
        "rows": rows,
        "warnings": warnings,
        **extra,
    }


def preview_manifest(content: bytes, competition_id: str) -> dict:
    parsed = read_manifest(content, competition_id)
    return _summarize(parsed["rows"], parsed["warnings"], sheet=parsed.get("sheet"))


def preview_rooming(content: bytes, competition_id: str) -> dict:
    parsed = read_rooming(content, competition_id)
    return _summarize(
        parsed["rows"], parsed["warnings"],
        sheet=parsed.get("sheet"),
        dates=parsed.get("dates", []),
        total_nights=parsed.get("total_nights", 0),
    )


# ── Commit (escribe) ─────────────────────────────────────────────────────────

def _existing_participants(competition_id: str) -> list[dict]:
    return (
        supabase.table("logistics_participants")
        .select("id,first_name,last_name,personnel_id,employee_id")
        .eq("competition_id", competition_id)
        .execute()
        .data
    ) or []


def _resolve_participant(row: dict, existing: list[dict], competition_id: str,
                         warnings: Optional[list] = None) -> str:
    """Id del participante de esta fila: el que ya está, o uno nuevo.

    Reusar en vez de insertar es lo que permite correr el manifest y la rooming
    list uno después del otro sin duplicar a nadie.
    """
    match = row["match"]
    if match["status"] == "matched" and match["id"]:
        key = "personnel_id" if match["kind"] == "personnel" else "employee_id"
        for p in existing:
            if p.get(key) == match["id"]:
                return p["id"]

    incoming = " ".join(x for x in (row.get("first_name"), row.get("last_name")) if x)
    twin = None
    for p in existing:
        current = " ".join(x for x in (p.get("first_name"), p.get("last_name")) if x)
        if _similarity(incoming, current) >= _MATCH_THRESHOLD:
            return p["id"]
        if twin is None and _possible_duplicate(incoming, current):
            twin = current

    # Se parece a alguien que ya está, pero no lo suficiente como para
    # fusionarlos sin preguntar. Un typo en un apellido corto ("Juyo"/"Guyo")
    # es indistinguible de dos apellidos distintos, así que en vez de adivinar
    # se crea la fila y se avisa: duplicar es molesto, fusionar mal es peor.
    if warnings is not None and twin:
        warnings.append({
            "row": row.get("source_row", 0),
            "field": "duplicate",
            "message": f"'{incoming}' podría ser la misma persona que '{twin}', que ya está "
                       f"en el padrón — se agrega aparte; revísalo",
        })

    record = {
        "competition_id": competition_id,
        "first_name": row.get("first_name") or "?",
        "last_name": row.get("last_name"),
        "role": row.get("role"),
        "category": row.get("category") or "official",
        "pax": row.get("pax", 1),
    }
    if row.get("passport"):
        record["passport"] = row["passport"]
    if row.get("dietary_restrictions"):
        record["dietary_restrictions"] = row["dietary_restrictions"]
    if match["status"] == "matched" and match["id"]:
        record["personnel_id" if match["kind"] == "personnel" else "employee_id"] = match["id"]

    created = supabase.table("logistics_participants").insert(record).execute().data[0]
    existing.append({
        "id": created["id"],
        "first_name": created.get("first_name"),
        "last_name": created.get("last_name"),
        "personnel_id": created.get("personnel_id"),
        "employee_id": created.get("employee_id"),
    })
    return created["id"]


def commit_manifest(content: bytes, competition_id: str) -> dict:
    parsed = read_manifest(content, competition_id)
    existing = _existing_participants(competition_id)
    warnings = list(parsed["warnings"])

    created_people = len(existing)
    legs_written = 0
    for row in parsed["rows"]:
        participant_id = _resolve_participant(row, existing, competition_id, warnings)
        if not row["legs"]:
            continue
        # Reemplazo, no upsert: la planilla es la autoridad sobre los vuelos, y
        # reimportarla no debe dejar tramos viejos dando vueltas.
        supabase.table("logistics_travel_legs").delete().eq("participant_id", participant_id).execute()
        supabase.table("logistics_travel_legs").insert(
            [{**leg, "participant_id": participant_id} for leg in row["legs"]]
        ).execute()
        legs_written += len(row["legs"])

    return {
        "participants_total": len(parsed["rows"]),
        "participants_created": len(existing) - created_people,
        "legs": legs_written,
        "warnings": warnings,
    }


def commit_rooming(content: bytes, competition_id: str) -> dict:
    parsed = read_rooming(content, competition_id)
    existing = _existing_participants(competition_id)
    warnings = list(parsed["warnings"])
    before = len(existing)

    # Hotel por defecto: la rooming list no nombra el hotel, así que se cuelga
    # del principal de la competencia si hay uno cargado.
    hotels = (
        supabase.table("logistics_hotels")
        .select("id,is_primary")
        .eq("competition_id", competition_id)
        .execute()
        .data
    ) or []
    hotel_id = next((h["id"] for h in hotels if h.get("is_primary")), None)
    if hotel_id is None and hotels:
        hotel_id = hotels[0]["id"]

    row_to_participant: dict = {}
    stays_written = 0
    for row in parsed["rows"]:
        participant_id = _resolve_participant(row, existing, competition_id, warnings)
        row_to_participant[row["source_row"]] = participant_id

        if row.get("dietary_restrictions"):
            supabase.table("logistics_participants").update(
                {"dietary_restrictions": row["dietary_restrictions"]}
            ).eq("id", participant_id).execute()

        stay = row["stay"]
        record = {
            "participant_id": participant_id,
            "hotel_id": hotel_id,
            "room_type": stay["room_type"],
            "room_number": stay["room_number"],
            "check_in": stay["check_in"],
            "check_out": stay["check_out"],
            "comments": stay["comments"],
        }
        supabase.table("logistics_stays").delete().eq("participant_id", participant_id).execute()
        supabase.table("logistics_stays").insert(record).execute()
        stays_written += 1

    # Las parejas de habitación se resuelven en una segunda pasada: recién acá
    # existen los ids de las dos personas.
    for row in parsed["rows"]:
        partner_row = row["stay"].get("share_with_row")
        if not partner_row:
            continue
        me = row_to_participant.get(row["source_row"])
        partner = row_to_participant.get(partner_row)
        if not me or not partner:
            continue
        supabase.table("logistics_stays").update({"share_with_id": partner}).eq("participant_id", me).execute()
        supabase.table("logistics_stays").update({"share_with_id": me}).eq("participant_id", partner).execute()

    # Habitaciones de 3+ personas: share_with_id es una FK singular (pairwise),
    # así que un cuarto triple/cuádruple no se puede representar completo —el
    # holder termina apuntando a una sola compañera y las demás se pierden en
    # silencio, incluido en el conteo de noches que se factura al hotel—. En vez
    # de tragarlo, detectar la componente conexa de "comparte con" y avisar.
    _share_adj: dict = {}
    for row in parsed["rows"]:
        partner_row = row["stay"].get("share_with_row")
        if not partner_row:
            continue
        a, b = row["source_row"], partner_row
        _share_adj.setdefault(a, set()).add(b)
        _share_adj.setdefault(b, set()).add(a)
    _seen_rows: set = set()
    for start in _share_adj:
        if start in _seen_rows:
            continue
        stack, comp = [start], set()
        while stack:
            node = stack.pop()
            if node in comp:
                continue
            comp.add(node)
            _seen_rows.add(node)
            stack.extend(_share_adj.get(node, ()))
        if len(comp) > 2:
            warnings.append({
                "row": min(comp),
                "field": "share_with",
                "message": (
                    f"Habitación compartida por {len(comp)} personas "
                    f"(filas {sorted(comp)}): el modelo solo guarda parejas, "
                    f"revisá esta habitación en el rooming a mano."
                ),
            })

    return {
        "participants_total": len(parsed["rows"]),
        "participants_created": len(existing) - before,
        "stays": stays_written,
        "total_nights": parsed.get("total_nights", 0),
        "warnings": warnings,
    }


# ── Export ───────────────────────────────────────────────────────────────────

def _autosize(ws) -> None:
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 40)


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_manifest_xlsx(competition_id: str) -> bytes:
    """Manifest en el mismo layout que usa la organización: llegadas a la
    izquierda, salidas a la derecha, una fila por persona."""
    from api._lib.routers.logistics import PARTICIPANT_FIELDS, full_name

    people = (
        supabase.table("logistics_participants")
        .select(PARTICIPANT_FIELDS)
        .eq("competition_id", competition_id)
        .order("last_name")
        .execute()
        .data
    ) or []
    legs_by_person: dict = {}
    if people:
        legs = (
            supabase.table("logistics_travel_legs")
            .select("*")
            .in_("participant_id", [p["id"] for p in people])
            .execute()
            .data
        ) or []
        for leg in legs:
            legs_by_person.setdefault(leg["participant_id"], []).append(leg)

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"
    ws.append([
        "#", "Name", "Last Name", "Role", "Passport #",
        "Arrival Date", "Arrival Time", "Flight #", "Arriving Airport", "Comments",
        "Departure Date", "Pick Up Time", "Departure Flight", "Flight #",
        "Departing Airport", "Comments",
    ])
    for idx, p in enumerate(people, start=1):
        person_legs = legs_by_person.get(p["id"], [])
        arrival = next((l for l in person_legs if l["direction"] == "arrival"), {}) or {}
        departure = next((l for l in person_legs if l["direction"] == "departure"), {}) or {}
        ws.append([
            idx, p.get("first_name"), p.get("last_name"), p.get("role"), p.get("passport"),
            arrival.get("date"), arrival.get("time"), arrival.get("flight_number"),
            arrival.get("airport"), arrival.get("comments"),
            departure.get("date"), departure.get("pickup_time"), departure.get("time"),
            departure.get("flight_number"), departure.get("airport"), departure.get("comments"),
        ])
    _autosize(ws)
    return _save(wb)


def export_rooming_xlsx(competition_id: str) -> bytes:
    """Rooming list con la grilla de noches y el total, que es lo que espera
    recibir el hotel."""
    from api._lib.routers.logistics import build_rooming

    data = build_rooming(competition_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rooming"

    header = ["#", "Name", "Last Name", "Role", "Room", "Share with",
              "Dietary Restrictions"] + data["dates"] + ["TOTAL", "Comments"]
    ws.append(header)

    for idx, row in enumerate(data["rows"], start=1):
        grid = [row["grid"].get(d, "") for d in data["dates"]]
        ws.append([
            idx, row.get("first_name"), row.get("last_name"), row.get("role"),
            row.get("room_type"), row.get("share_with"), row.get("dietary_restrictions"),
            *grid, row["nights"], row.get("comments"),
        ])

    ws.append([""] * (len(header) - 2) + [data["total_nights"], ""])
    _autosize(ws)
    return _save(wb)
