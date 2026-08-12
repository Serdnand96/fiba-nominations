"""Import de presupuesto (`budget_lines`) desde Excel.

Reemplaza en la UI lo que hasta ahora era un script de una sola vez
(`scripts/import_budgets_2027.py`): cargar un presupuesto ya armado en Excel
sin picar 168 celdas a mano.

Dos pasos como el resto de los importadores del repo (preview → commit) y
**nunca adivina plata**: una columna que no matchea una competencia se reporta
con su monto y se deja afuera; no se manda a "General", que es una columna real
del Excel y ensuciarla arruinaría el reporte. El usuario resuelve las dudosas a
mano en el preview y esa decisión viaja al commit.

Las dos formas reales que se soportan salen de las dos planillas del cliente
(ver BUDGET_MODULE.md §1):

  * **Lista** — una línea por fila, con código contable y, opcionalmente, una
    columna por año (`Annual 2027`, `Annual 2028`, …). Es el presupuesto de IT.
    Las filas de un mismo concepto a través de los años comparten `series_id`.
  * **Matriz** — filas = líneas de gasto, columnas = competencias, más una
    columna "General". Es el presupuesto de Competitions.

La forma se detecta por los encabezados, no por posición: las planillas reales
tienen títulos arriba, celdas combinadas, saltos de línea dentro del header y
filas de subtotal en el medio.

Idempotente sin marcas: una fila que matchea una línea ya cargada
(año + departamento + cuenta + competencia + descripción) la **actualiza** en vez
de duplicarla. Reimportar la planilla corregida es la operación normal.
"""
import io
import re
import unicodedata
import uuid
from difflib import SequenceMatcher
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from api._lib.database import supabase

# Claves especiales del mapeo de columnas/valores de competencia.
GENERAL = "__general__"   # gasto sin evento (competition_id = NULL)
SKIP = "__skip__"         # no importar esta columna

MAX_ROWS = 2000           # filas de datos por planilla
MAX_HEADER_SCAN = 25      # filas donde buscar el encabezado

_MATCH_THRESHOLD = 0.92   # arriba de esto se vincula la competencia sola
_REVIEW_THRESHOLD = 0.60  # entre medio, se sugiere pero NO se aplica

# Caracteres con los que un texto exportado podría interpretarse como fórmula
# al abrirlo (CWE-1236). Mismo criterio que schedule_workbook.
_FORMULA_LEAD = ("=", "+", "-", "@")


# ─── Normalización ───────────────────────────────────────────────────────────
def norm(value) -> str:
    """minúsculas, sin acentos, sin saltos de línea, espacios colapsados."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip(" :*")


def clean(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text) or None


_NULLISH = {"", "-", "--", "—", "n/a", "na", "s/d", "tbc", "tbd"}


def to_number(value) -> Optional[float]:
    """Número de una celda, tolerando '$ 1.234,50', '(500)' y guiones.

    Las planillas vienen tipeadas a mano: hay montos como texto, con símbolo de
    moneda y con separadores en cualquiera de los dos formatos. Un monto mal
    leído es plata mal presupuestada, así que lo que no se puede interpretar
    devuelve None (y la fila se reporta) en vez de caer a 0.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if norm(text) in _NULLISH:
        return None
    negative = text.startswith("(") and text.endswith(")")
    # Solo se sacan moneda, espacios y el signo de porcentaje. Las letras NO se
    # descartan: si se descartaran, "All Star 5 Gifts" valdría 5 y un concepto
    # entraría como monto.
    text = re.sub(r"(?i)\b(usd|us\$|eur)\b", "", text.strip("()").strip())
    text = re.sub(r"[\s $€%]", "", text)
    if not re.fullmatch(r"[+-]?[\d.,]+", text) or not re.search(r"\d", text):
        return None
    if "," in text and "." in text:
        # el separador decimal es el último que aparece
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif "," in text:
        # "1,50" es decimal; "1,500" y "1,234,567" son miles
        text = text.replace(",", ".") if re.fullmatch(r"-?\d+,\d{1,2}", text) else text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative and number > 0 else number


_CONTAINMENT_CAP = 0.85   # < _MATCH_THRESHOLD, a propósito (ver abajo)


def _words(text: str) -> set:
    """Palabras significativas, sin puntuación.

    Los paréntesis y guiones separan lo mismo escrito de dos formas —
    "Long Term Acquisitions (Court / Backstop Units)" y
    "Long Term Acquisitions - Court / Backstop Units"— así que se descartan.
    """
    return {w for w in re.split(r"[^a-z0-9]+", norm(text)) if len(w) > 1}


def _similarity(a: str, b: str) -> float:
    """Parecido entre dos nombres de competencia, sin importar el orden.

    ⚠️ La **contención** (un nombre corto que entra entero en uno largo) está
    topeada por debajo del umbral de auto-match: "Women's AmeriCup" está
    contenido en "FIBA U16 Women's AmeriCup" y son dos eventos distintos con
    presupuestos distintos. Sirve para *sugerir* la columna, nunca para
    vincularla sola. Es la diferencia con el matcher de personas de Logística,
    donde la planilla sí acorta el nombre legal de la misma persona.
    """
    words_a, words_b = _words(a), _words(b)
    if not words_a or not words_b:
        return 0.0
    shared = len(words_a & words_b)
    containment = shared / min(len(words_a), len(words_b))
    fuzzy = SequenceMatcher(None, " ".join(sorted(words_a)), " ".join(sorted(words_b))).ratio()
    return max(fuzzy, min(containment, _CONTAINMENT_CAP))


# ─── Clasificación de encabezados ────────────────────────────────────────────
#
# El orden importa: se prueba coincidencia exacta, después año, después
# subcadena. "Monthly 2027" es la columna mensual y no el año 2027; "General
# Expenses (not associated to the competitions)" es la columna general y no una
# columna de competencia, aunque diga "competitions".

_EXACT = {
    "id": "id", "line id": "id", "budget line id": "id",
    "competition": "competition", "competencia": "competition",
    "event": "competition", "evento": "competition", "competition name": "competition",
    "qt": "qty", "qty": "qty", "cant": "qty",
    "code": "account", "cta": "account",
    "ref": "ignore", "line": "ignore", "linea": "ignore", "#": "ignore",
    "no.": "ignore", "nro": "ignore", "n°": "ignore",
}

_SUBSTRING = [
    ("account_label", ("account name", "account label", "nombre de cuenta", "nombre de la cuenta")),
    ("account", ("account code", "account", "cuenta", "codigo contable", "codigo")),
    ("department", ("department", "departamento")),
    ("qty", ("quantity", "cantidad")),
    ("monthly", ("monthly", "mensual")),
    ("escalation", ("escalation", "escalacion", "incremento")),
    ("notes", ("notes", "notas", "comentario", "comment", "observacion", "explanation", "explicacion", "clarification")),
    ("description", ("description", "descripcion", "detalle", "expense item", "concepto", "line item", "detail", "item")),
    ("amount", ("amount", "monto", "importe", "total", "annual", "anual", "budget", "presupuesto")),
]

_YEAR_RE = re.compile(r"\b(20\d{2})\b")
# palabras que pueden acompañar a un año sin dejar de ser una columna de año
_YEAR_WORDS = {"annual", "anual", "budget", "presupuesto", "fy", "year", "ano", "a",
               "total", "est", "estimated", "estimado", "forecast", "proyectado",
               "proyeccion", "monto", "amount", "usd"}


def _year_of(header: str) -> Optional[int]:
    match = _YEAR_RE.search(header)
    if not match:
        return None
    rest = [w for w in re.split(r"[^a-z]+", _YEAR_RE.sub(" ", header)) if w]
    if all(w in _YEAR_WORDS for w in rest):
        year = int(match.group(1))
        if 2000 <= year <= 2100:
            return year
    return None


def _classify(header) -> tuple:
    """(role, year) para un encabezado. role='unknown' = candidata a competencia."""
    text = norm(header)
    if not text:
        return ("blank", None)
    if text in _EXACT:
        return (_EXACT[text], None)
    year = _year_of(text)
    if year:
        return ("year", year)
    if "general" in text:
        return ("general", None)
    for role, keys in _SUBSTRING:
        if any(key in text for key in keys):
            return (role, None)
    return ("unknown", None)


# ─── Lectura de la planilla (sin tocar la base) ──────────────────────────────
def _sheet_rows(ws) -> list:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _header_score(cells: list) -> tuple:
    """(puntaje, roles) de una fila candidata a encabezado.

    Sin descripción no hay línea que cargar, pero con la descripción sola no
    alcanza: "Game Experience-Promotional Items" es una fila de datos que
    contiene la palabra "item". Hace falta además otra columna reconocida (el
    caso lista) o varias columnas de texto sin rol (el caso matriz, donde cada
    una es una competencia).
    """
    roles = [_classify(c) for c in cells]
    if not any(r == "description" for r, _ in roles):
        return (0, roles)
    known = {r for r, _ in roles} - {"blank", "unknown", "ignore", "description"}
    unknown_text = sum(
        1 for i, (role, _) in enumerate(roles)
        if role == "unknown" and clean(cells[i]) and to_number(cells[i]) is None
    )
    if not known and unknown_text < 2:
        return (0, roles)
    return (len(known) + 3, roles)


def _analyze_sheet(ws) -> dict:
    """Encabezado, columnas y filas crudas de una hoja. Sin base de datos."""
    rows = _sheet_rows(ws)
    best_idx, best_score, best_roles = None, 0, []
    for idx, cells in enumerate(rows[:MAX_HEADER_SCAN]):
        score, roles = _header_score(cells)
        if score > best_score:
            best_idx, best_score, best_roles = idx, score, roles

    if best_idx is None:
        return _analyze_headerless(rows)

    columns = []
    for i, (role, year) in enumerate(best_roles):
        header = clean(rows[best_idx][i]) if i < len(rows[best_idx]) else None
        columns.append({"index": i, "header": header, "role": role, "year": year})
    return {"header_row": best_idx + 1, "columns": columns,
            "rows": rows[best_idx + 1:], "headerless": False,
            "data_start": best_idx + 2}


def _analyze_headerless(rows: list) -> dict:
    """Planilla sin encabezados: una columna de concepto y una de monto.

    Es la forma de los presupuestos armados a mano: el de un solo evento
    ("AmeriCup W 2027") y el de Comms, que arranca con un título y ni siquiera
    usa la columna A. Se busca el par de columnas —cualquier par, no A|B— donde
    el patrón texto+número se repite tres veces seguidas. Sin esa repetición se
    rechaza: adivinar qué columna es cuál sería inventar el presupuesto.
    """
    candidates = []
    width = max((len(r) for r in rows), default=0)
    for text_col in range(min(width, 8)):
        for number_col in range(text_col + 1, min(width, 12)):
            first, streak, hits, best_streak, values = None, 0, 0, 0, []
            for idx, cells in enumerate(rows):
                text_ok = len(cells) > text_col and bool(clean(cells[text_col])) \
                    and to_number(cells[text_col]) is None
                number = to_number(cells[number_col]) if len(cells) > number_col else None
                if text_ok and number is not None:
                    if first is None:
                        first = idx
                    streak += 1
                    hits += 1
                    best_streak = max(best_streak, streak)
                    values.append(abs(number))
                else:
                    streak = 0
            # Una columna de puros valores < 1 es un porcentaje, no plata: el
            # presupuesto de Comms trae el % de cada línea al lado del monto.
            fraction = bool(values) and all(v < 1 for v in values)
            if best_streak >= 3 and first is not None:
                candidates.append({"hits": hits, "text": text_col, "number": number_col,
                                   "first": first, "fraction": fraction})
    if not candidates:
        return {"header_row": None, "columns": [], "rows": [], "headerless": True, "data_start": 1}

    money_cols = [c for c in candidates if not c["fraction"]] or candidates
    top = max(c["hits"] for c in money_cols)
    # A igualdad de cobertura gana la columna de más a la izquierda: es la que
    # está pegada al concepto.
    best = min((c for c in money_cols if c["hits"] >= 0.9 * top),
               key=lambda c: (c["number"], c["text"]))
    text_col, number_col, first = best["text"], best["number"], best["first"]
    columns = [{"index": text_col, "header": None, "role": "description", "year": None},
               {"index": number_col, "header": None, "role": "amount", "year": None}]

    # Una columna de texto libre a la derecha del monto suele ser la de notas
    # ("NOTES" en el de Comms): es información que alguien escribió a mano y
    # tirarla sería peor que arriesgar una columna mal rotulada, que se ve.
    data = rows[first:]
    for index in range(number_col + 1, min(width, 12)):
        texts = sum(1 for r in data if len(r) > index and clean(r[index]) and to_number(r[index]) is None)
        if texts >= max(3, len(data) // 4):
            columns.append({"index": index, "header": None, "role": "notes", "year": None})
            break

    return {"header_row": None, "columns": columns, "rows": data, "headerless": True,
            "data_start": first + 1}


def _detect_groups(entries: list, candidates: set) -> tuple:
    """Detecta los subtotales de una lista jerárquica. → (grupos, padre_de)

    El presupuesto de Comms viene como árbol: una fila por evento con su
    subtotal ("BCLA 100.000") y debajo sus líneas ("Photo Ops 24.000", …),
    anidado dos niveles ("EVENTS" contiene cuatro eventos). Importar todo
    plano contaría la plata dos veces.

    Dos señales, y hacen falta las dos:

    * **candidata** — la fila viene en negrita o menos sangrada que la de
      abajo. Es lo que separa un subtotal de una línea cualquiera.
    * **aritmética** — su monto es exactamente la suma de las filas que le
      siguen y todavía no pertenecen a otro grupo.

    Con la aritmética sola no alcanza: en esta misma planilla "Draw 8.000"
    coincide con las dos líneas de abajo (5.000 + 3.000) y se robaba los hijos
    de BCLA, que entonces dejaba de cerrar. Con la negrita sola tampoco: si las
    sumas no cierran no hay árbol y la planilla se lee plana.

    Se recorre de abajo hacia arriba, así los grupos internos quedan resueltos
    antes que los que los contienen.

    `entries` es [(offset, descripción, monto)] en el orden de la planilla.
    """
    groups, parent_of = set(), {}
    total = len(entries)
    for i in range(total - 1, -1, -1):
        offset, _, amount = entries[i]
        if amount is None or amount <= 0 or offset not in candidates:
            continue
        running, children = 0.0, []
        for j in range(i + 1, total):
            if entries[j][0] in parent_of:
                continue                      # ya es hijo de un grupo interno
            child_amount = entries[j][2]
            if child_amount is None:
                continue
            running += child_amount
            children.append(entries[j][0])
            if round(running, 2) >= round(amount, 2):
                break
        # Dos hijos como mínimo: que una línea suelta coincida con la de abajo
        # es una casualidad, no un subtotal.
        if len(children) >= 2 and round(running, 2) == round(amount, 2):
            groups.add(offset)
            for child in children:
                parent_of[child] = offset
    return groups, parent_of


def _shape(columns: list, rows: list) -> str:
    """'matrix' si hay ≥3 columnas sin rol con números; si no 'list'.

    Tres y no dos: las planillas de seguimiento traen pares "Rate | Total"
    (presupuesto y actuals) que sin el encabezado agrupador de arriba son
    indistinguibles de dos eventos. Con dos columnas sueltas conviene leer la
    planilla como lista y decir qué columna se usó, que es visible y corregible;
    leerla como matriz inventaría dos competencias.
    """
    if any(c["role"] == "year" for c in columns):
        return "list"
    candidates = [c for c in columns if c["role"] in ("unknown", "general")]
    with_numbers = [
        c for c in candidates
        if any(to_number(r[c["index"]]) is not None for r in rows if len(r) > c["index"])
    ]
    return "matrix" if len(with_numbers) >= 3 else "list"


def _has_formulas(content: bytes, sheet: str) -> bool:
    """¿El archivo trae fórmulas sin valor calculado?

    openpyxl con data_only=True devuelve None cuando la planilla se generó con
    una herramienta que no cachea los resultados. Sin este chequeo el import
    saldría "0 filas" sin explicar por qué.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.startswith("="):
                    return True
    except Exception:
        return False
    return False


def _emphasized_rows(ws, parsed: dict) -> set:
    """Offsets de las filas que *parecen* un encabezado de grupo.

    Negrita, o menos sangría que la fila siguiente. Es solo la lista de
    candidatas de `_detect_groups`: que una fila esté en negrita no la
    convierte en subtotal, lo decide la aritmética.
    """
    column = next((c for c in parsed["columns"] if c["role"] == "description"), None)
    if not column:
        return set()
    start = parsed.get("data_start", 1)
    indents, bold = {}, set()
    for offset in range(len(parsed["rows"])):
        try:
            cell = ws.cell(row=start + offset, column=column["index"] + 1)
        except (ValueError, IndexError):
            continue
        if cell.font is not None and cell.font.bold:
            bold.add(offset)
        indents[offset] = int((cell.alignment.indent if cell.alignment else 0) or 0)

    emphasized = set(bold)
    for offset, indent in indents.items():
        following = indents.get(offset + 1)
        if following is not None and following > indent:
            emphasized.add(offset)
    return emphasized


def analyze(content: bytes, sheet: Optional[str] = None) -> dict:
    """Lee la planilla y devuelve columnas + filas crudas. Sin base de datos."""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheets = list(wb.sheetnames)
    if sheet and sheet not in sheets:
        raise ValueError(f"Sheet '{sheet}' not found")

    if sheet:
        chosen = sheet
        parsed = _analyze_sheet(wb[sheet])
    else:
        # Auto-elección: mejor encabezado y, a igualdad, más celdas con datos.
        # El workbook de Competitions trae "Summary" (condensado) y "Staffing
        # (travel)" (headcount, no plata) con la misma forma que el Breakdown.
        best = None
        for name in sheets:
            candidate = _analyze_sheet(wb[name])
            score = _header_score(_sheet_rows(wb[name])[candidate["header_row"] - 1])[0] if candidate["header_row"] else 1
            cells = sum(
                1 for r in candidate["rows"] for c in candidate["columns"]
                if len(r) > c["index"] and to_number(r[c["index"]]) is not None
            )
            if candidate["columns"] and (best is None or (score, cells) > best[0]):
                best = ((score, cells), name, candidate)
        if best is None:
            raise ValueError(
                "No pude reconocer el formato: ninguna hoja tiene una columna de "
                "descripción. Descarga la plantilla desde el botón de al lado."
            )
        _, chosen, parsed = best

    if not parsed["columns"]:
        raise ValueError(
            f"No pude reconocer el formato de la hoja '{chosen}'. Necesito al menos "
            "una columna de descripción y una de monto (o una columna por competencia)."
        )

    parsed["sheet"] = chosen
    parsed["sheets"] = sheets
    parsed["shape"] = _shape(parsed["columns"], parsed["rows"])
    parsed["emphasized"] = _emphasized_rows(wb[chosen], parsed)
    if parsed["shape"] == "matrix":
        # En matriz la columna "Total"/"TOTALS" es el total de la fila: sumarla
        # duplicaría el presupuesto entero.
        for column in parsed["columns"]:
            if column["role"] == "amount":
                column["role"] = "row_total"
    return parsed


# ─── Catálogos de la base ────────────────────────────────────────────────────
class Catalog:
    """Competencias, cuentas y departamentos contra los que se resuelve."""

    def __init__(self, year: int, kind: str, viewable: Optional[list]):
        self.year = year
        self.kind = kind
        self.competitions = supabase.table("competitions").select("id, name, year").execute().data or []
        self.accounts = supabase.table("accounts").select("code, label, kind, active").execute().data or []
        self.departments = supabase.table("departments").select("code, label, active").execute().data or []
        self._by_label = {}
        for account in self.accounts:
            if account.get("kind") == kind and account.get("active"):
                self._by_label.setdefault(norm(account.get("label")), account["code"])
        self._accounts_by_code = {a["code"]: a for a in self.accounts}
        self._dept_codes = {d["code"] for d in self.departments if d.get("active")}
        self._dept_by_label = {norm(d["label"]): d["code"] for d in self.departments if d.get("active")}
        self.viewable = viewable

    # ── competencias ────────────────────────────────────────────────────────
    def match_competition(self, label: str) -> dict:
        """{status, competition_id, name, score} para un nombre de la planilla.

        Solo un match casi exacto se aplica solo. Lo dudoso se devuelve como
        sugerencia sin aplicar: acá una confusión no es un typo en un nombre,
        es plata imputada a otro evento.
        """
        text = norm(label)
        if not text:
            return {"status": "unmatched", "competition_id": None, "name": None, "score": 0}
        pool = [c for c in self.competitions if c.get("year") == self.year] or self.competitions
        best, best_score = None, 0.0
        for competition in pool:
            if norm(competition["name"]) == text:
                return {"status": "matched", "competition_id": competition["id"],
                        "name": competition["name"], "score": 1.0}
            score = _similarity(label, competition["name"])
            if score > best_score:
                best, best_score = competition, score
        if best and best_score >= _MATCH_THRESHOLD:
            return {"status": "matched", "competition_id": best["id"],
                    "name": best["name"], "score": round(best_score, 2)}
        if best and best_score >= _REVIEW_THRESHOLD:
            return {"status": "review", "competition_id": None,
                    "name": best["name"], "score": round(best_score, 2),
                    "suggested_id": best["id"]}
        return {"status": "unmatched", "competition_id": None, "name": None, "score": 0}

    # ── cuentas ─────────────────────────────────────────────────────────────
    def account_by_code(self, code: str) -> Optional[dict]:
        return self._accounts_by_code.get(str(code).strip())

    def account_by_label(self, label: str) -> tuple:
        """(cuenta, score) para el nombre de una línea sin código contable.

        La planilla de Competitions no trae códigos: la cuenta es el propio
        "Expense Item". El seed guardó esos textos como label de `COMP-01`…`28`,
        pero limpiados ("TV Production / Streaming" por "TV Production /
        Streaming (COMMS)"), así que el match exacto solo no alcanza y
        reimportar la misma planilla crearía cuentas duplicadas.

        Acá la contención sí sirve —la cuenta es una categoría, no un evento, y
        el ítem de la planilla suele ser la cuenta más un calificativo—, pero
        exige dos palabras en común: "Shipping" no se come "Shipping (DHL or
        Sea)". Todo match no exacto se reporta en el preview.
        """
        text = norm(label)
        if not text:
            return (None, 0.0)
        if text in self._by_label:
            code = self._by_label[text]
            return (self._accounts_by_code.get(code), 1.0)

        words = _words(text)
        best, best_score = None, 0.0
        for account in self.accounts:
            if account.get("kind") != self.kind or not account.get("active"):
                continue
            other = _words(account.get("label"))
            if not words or not other:
                continue
            shared = len(words & other)
            contained = shared >= 2 and shared == min(len(words), len(other))
            score = 0.95 if contained else _similarity(text, account.get("label"))
            if score > best_score:
                best, best_score = account, score
        return (best, round(best_score, 2)) if best_score >= 0.90 else (None, 0.0)

    def next_provisional_code(self, department: str, taken: set) -> str:
        """Código provisorio `<DEPT>-NN` para una cuenta que no existe todavía.

        El plan de cuentas oficial de Finance no llegó (BUDGET_MODULE.md §12):
        las 28 líneas de Competitions ya viven como `COMP-01`…`COMP-28` con
        `pending_mapping = true`. Las nuevas siguen esa misma convención en vez
        de inventar un código contable que después habría que desarmar.
        """
        # 4 letras: el departamento `competitions` da COMP-29 y sigue la serie
        # COMP-01…COMP-28 que ya existe en vez de abrir una convención nueva.
        prefix = re.sub(r"[^A-Z0-9]", "", (department or "gen").upper())[:4]
        used = {c for c in self._accounts_by_code} | taken
        index = 1
        while f"{prefix}-{index:02d}" in used:
            index += 1
        return f"{prefix}-{index:02d}"

    # ── departamentos ───────────────────────────────────────────────────────
    def department_code(self, value) -> Optional[str]:
        text = norm(value)
        if not text:
            return None
        if text in self._dept_codes:
            return text
        return self._dept_by_label.get(text)

    # ── líneas ya cargadas ──────────────────────────────────────────────────
    def existing_lines(self, years: list) -> list:
        query = (
            supabase.table("budget_lines")
            .select("id, year, department_code, account_code, competition_id, description, series_id, kind")
            .in_("year", sorted(set(years)))
        )
        if self.viewable is not None:
            query = query.in_("department_code", self.viewable)
        return [r for r in (query.execute().data or []) if r.get("kind", "expense") == self.kind]


# ─── Armado del plan ─────────────────────────────────────────────────────────
def _row_key(year, department, account, competition_id, description) -> tuple:
    return (year, department, account, competition_id or None, norm(description))


def _plan(content: bytes, *, year: int, kind: str, department: str,
          sheet: Optional[str] = None, mapping: Optional[dict] = None,
          default_competition_id: Optional[str] = None,
          create_accounts: bool = True, replace: bool = False,
          editable: Optional[list] = None, viewable: Optional[list] = None) -> dict:
    """Todo el trabajo: leer, resolver contra la base y decidir qué haría cada fila.

    `preview` lo muestra y `commit` lo ejecuta — la misma función, para que lo
    que se confirma sea exactamente lo que se vio.
    """
    parsed = analyze(content, sheet)
    columns, raw_rows, shape = parsed["columns"], parsed["rows"], parsed["shape"]
    mapping = mapping or {}
    warnings, errors = [], []

    def column_of(*roles):
        for column in columns:
            if column["role"] in roles:
                return column
        return None

    col_desc = column_of("description")
    col_account = column_of("account")
    col_account_label = column_of("account_label")
    col_department = column_of("department")
    col_qty = column_of("qty")
    col_monthly = column_of("monthly")
    col_escalation = column_of("escalation")
    col_notes = column_of("notes")
    col_id = column_of("id")
    col_competition = column_of("competition")
    year_columns = [c for c in columns if c["role"] == "year"]
    comp_columns = [c for c in columns if c["role"] in ("unknown", "general")] if shape == "matrix" else []
    col_amount = None
    if shape == "list":
        amounts = [c for c in columns if c["role"] == "amount"]
        # La primera de izquierda a derecha: cuando hay dos ("Total" del
        # presupuesto y "Total" de actuals) la del presupuesto viene antes.
        col_amount = amounts[0] if amounts else None
        if not year_columns and not col_amount:
            raise ValueError(
                "No encontré la columna de monto. Debe llamarse 'Monto'/'Amount' "
                "o haber una columna por año ('Annual 2027')."
            )
        if not year_columns and len(amounts) > 1:
            warnings.append({"row": parsed["header_row"], "message":
                             f"La planilla tiene {len(amounts)} columnas de monto; se usó «{col_amount['header']}»"})

    years = sorted({c["year"] for c in year_columns}) or [year]
    catalog = Catalog(year, kind, viewable)

    # ── jerarquía (lista con subtotales por evento) ──────────────────────────
    # Antes de leer fila por fila hay que saber cuáles son subtotales: en el
    # presupuesto de Comms el evento es la fila de arriba ("BCLA 100.000") y sus
    # líneas cuelgan debajo. Sin esto la plata se contaría dos veces y el evento
    # se perdería.
    group_rows, group_labels = set(), {}
    if shape == "list" and not year_columns and not col_competition:
        entries, seen_any = [], False
        for offset, cells in enumerate(raw_rows):
            description = clean(cells[col_desc["index"]]) if len(cells) > col_desc["index"] else None
            if not description:
                continue
            head = norm(description)
            if head.startswith(("subtotal", "sub total", "sub-total")):
                continue
            entries.append((offset, description, to_number(cells[col_amount["index"]])
                            if col_amount and len(cells) > col_amount["index"] else None))
            seen_any = True
        if seen_any:
            detected, parent_of = _detect_groups(entries, parsed.get("emphasized") or set())
            by_offset = {e[0]: e[1] for e in entries}
            leaves = [e for e in entries if e[0] not in detected and (e[2] or 0) > 0]
            covered = [e for e in leaves if e[0] in parent_of]
            # Se acepta solo si el árbol explica la planilla: dos grupos o más y
            # la gran mayoría de las líneas colgando de alguno. Una coincidencia
            # aritmética suelta no alcanza para reinterpretar el archivo.
            if len(detected) >= 2 and leaves and len(covered) >= 0.6 * len(leaves):
                group_rows = detected
                group_labels = {e[0]: by_offset[parent_of[e[0]]] for e in covered}

    # ── resolución de las etiquetas de competencia ──────────────────────────
    # En matriz son los encabezados de columna; en lista, los valores de la
    # columna "Competencia". Se resuelven una sola vez y el usuario puede pisar
    # cualquiera desde el preview (`mapping`).
    labels = []
    if shape == "matrix":
        # Etiqueta única por columna: dos columnas pueden tener el mismo
        # encabezado (o ninguno) y el mapeo del usuario se guarda por etiqueta.
        seen_labels = set()
        for column in comp_columns:
            label = column["header"] or f"col {column['index'] + 1}"
            if label in seen_labels:
                label = f"{label} (col {column['index'] + 1})"
            seen_labels.add(label)
            column["label"] = label
            labels.append(label)
    elif col_competition:
        labels = sorted({clean(r[col_competition["index"]]) for r in raw_rows
                         if len(r) > col_competition["index"] and clean(r[col_competition["index"]])})
    elif group_labels:
        # Los subtotales de la jerarquía son la dimensión de competencia: se
        # resuelven con el mismo desplegable que las columnas de la matriz.
        labels = sorted(set(group_labels.values()))

    resolved = {}
    for label in labels:
        override = mapping.get(label)
        if override == SKIP:
            resolved[label] = {"status": "skipped", "competition_id": None, "name": None, "score": 0}
            continue
        if override == GENERAL:
            resolved[label] = {"status": "general", "competition_id": None, "name": None, "score": 0}
            continue
        if override:
            # Solo un id de competencia real: el mapeo llega del cliente y un
            # valor inventado terminaría como competition_id en la base.
            match = next((c for c in catalog.competitions if c["id"] == override), None)
            if not match:
                warnings.append({"row": None, "message": f"Columna «{label}»: la competencia elegida no existe"})
                resolved[label] = {"status": "skipped", "competition_id": None, "name": None, "score": 0}
                continue
            resolved[label] = {"status": "assigned", "competition_id": override,
                               "name": match["name"], "score": 1.0}
            continue
        if shape == "matrix" and any(c.get("label") == label and c["role"] == "general" for c in comp_columns):
            resolved[label] = {"status": "general", "competition_id": None, "name": None, "score": 0}
            continue
        resolved[label] = catalog.match_competition(label)

    # ── filas ───────────────────────────────────────────────────────────────
    existing = catalog.existing_lines(years)
    by_key = {}
    for line in existing:
        by_key.setdefault(_row_key(line["year"], line["department_code"], line["account_code"],
                                   line["competition_id"], line["description"]), line)
    by_id = {line["id"]: line for line in existing}
    # series existentes por concepto, para que los años queden unidos
    series_by_concept = {}
    for line in existing:
        series_by_concept.setdefault(
            (line["department_code"], line["account_code"], line["competition_id"], norm(line["description"])),
            line["series_id"],
        )

    items, skipped_cells = [], []
    new_accounts, seen_codes = {}, set()
    ignored_rows = 0
    stop = False

    for offset, cells in enumerate(raw_rows):
        if stop or len(items) > MAX_ROWS:
            break
        row_no = (parsed["header_row"] or 0) + offset + 1

        def cell(column):
            if not column or len(cells) <= column["index"]:
                return None
            return cells[column["index"]]

        description = clean(cell(col_desc))
        if not description:
            ignored_rows += 1
            continue
        head = norm(description)
        # "Total" / "TOTAL IT BUDGET" cierran la tabla: abajo solo quedan
        # totales, metadatos (SubZone, Tier) y notas. Una línea real que
        # empiece con "Total" (p. ej. "Total Station Rental") trae su código
        # de cuenta, así que no corta.
        if head.startswith(("total", "gran total", "grand total")) and items \
                and not clean(cell(col_account)):
            stop = True
            continue
        if head.startswith(("subtotal", "sub total", "sub-total")) or offset in group_rows:
            ignored_rows += 1     # subtotal: su plata ya está en las líneas de abajo
            continue

        # departamento: el de la fila si la planilla lo trae, si no el elegido.
        # La columna existe porque una competencia cruza departamentos: las
        # líneas de Comms e IT viven dentro del presupuesto de Competitions.
        row_department = department
        if col_department and clean(cell(col_department)):
            resolved_department = catalog.department_code(cell(col_department))
            if not resolved_department:
                warnings.append({"row": row_no, "message":
                                 f"Departamento '{clean(cell(col_department))}' desconocido — se usa {department}"})
            else:
                row_department = resolved_department

        qty = to_number(cell(col_qty))
        monthly = to_number(cell(col_monthly))
        escalation = to_number(cell(col_escalation))
        if escalation is not None:
            if escalation > 1:          # "5" o "5%" en vez de 0,05
                escalation = escalation / 100
            escalation = max(0.0, min(escalation, 9.9999))
        notes = clean(cell(col_notes))
        line_id = clean(cell(col_id))

        # (competencia, año, monto) de esta fila — una fila de la planilla puede
        # generar varias líneas: una por columna de competencia o por año.
        buckets = []
        if shape == "matrix":
            for column in comp_columns:
                amount = to_number(cells[column["index"]]) if len(cells) > column["index"] else None
                if amount is None or amount == 0:
                    continue
                label = column["label"]
                target = resolved.get(label, {})
                if target.get("status") in ("matched", "assigned", "general"):
                    buckets.append((target.get("competition_id"), year, amount, label))
                else:
                    skipped_cells.append({"row": row_no, "label": label, "amount": amount,
                                          "description": description})
        else:
            competition_id = default_competition_id
            label = clean(cell(col_competition)) if col_competition else group_labels.get(offset)
            if label:
                target = resolved.get(label, {})
                if target.get("status") in ("matched", "assigned"):
                    competition_id = target.get("competition_id")
                elif target.get("status") == "general":
                    competition_id = None
                else:
                    amount = to_number(cell(col_amount)) if col_amount else None
                    skipped_cells.append({"row": row_no, "label": label,
                                          "amount": amount or 0, "description": description})
                    continue
            if year_columns:
                for column in year_columns:
                    amount = to_number(cells[column["index"]]) if len(cells) > column["index"] else None
                    if amount is None or amount == 0:
                        continue
                    buckets.append((competition_id, column["year"], amount, label))
            else:
                amount = to_number(cell(col_amount))
                if amount is None or amount == 0:
                    # Una línea en cero no es presupuesto: las planillas dejan
                    # filas plantilla ("X", 0) al final del bloque.
                    ignored_rows += 1
                    continue
                buckets.append((competition_id, year, amount, label))

        if not buckets:
            ignored_rows += 1
            continue

        # ── cuenta ──────────────────────────────────────────────────────────
        # Se resuelve recién acá, con la fila ya confirmada como línea de
        # gasto: las planillas traen filas de agrupación ("610400 — Utilities")
        # sin montos, y resolverlas antes inventaba cuentas nuevas por cada
        # encabezado de grupo.
        account_code = clean(cell(col_account))
        account_label = clean(cell(col_account_label)) or description
        account_status = "existing"
        account_error = None
        if account_code:
            account = catalog.account_by_code(account_code)
            if not account:
                account_status = "new"
            elif account.get("kind") != kind:
                account_error = f"La cuenta {account_code} es de tipo {account.get('kind')}"
        else:
            matched_account, account_score = catalog.account_by_label(description)
            if matched_account:
                account_code = matched_account["code"]
                if account_score < 1.0:
                    warnings.append({"row": row_no, "message":
                                     f"«{description}» se imputa a {account_code} "
                                     f"«{matched_account['label']}» ({int(account_score * 100)}%)"})
            else:
                # Contra las cuentas nuevas de esta misma planilla, además del
                # texto exacto: el de Comms escribe "Photo Ops" en un evento y
                # "Photo Op" en el siguiente, y son la misma categoría. Sin
                # esto el import abre dos cuentas provisorias por cada plural.
                twin = next(
                    (v for k, v in new_accounts.items()
                     if k == norm(description) or _similarity(description, v["label"]) >= 0.92),
                    None,
                )
                account_code = (twin or {}).get("code") \
                    or catalog.next_provisional_code(row_department, seen_codes)
                account_status = "new"
        if account_status == "new":
            seen_codes.add(account_code)
            entry = new_accounts.setdefault(norm(description), {
                "code": account_code, "label": account_label[:120], "rows": 0, "amount": 0.0,
            })
            entry["rows"] += 1

        base_year = min(b[1] for b in buckets)
        concept = (row_department, account_code, buckets[0][0], norm(description))
        series_id = series_by_concept.get(concept) or str(uuid.uuid4())
        series_by_concept[concept] = series_id

        for competition_id, line_year, amount, label in buckets:
            if account_status == "new":
                new_accounts[norm(description)]["amount"] += amount
            item = {
                "row": row_no,
                "year": line_year,
                "department_code": row_department,
                "account_code": account_code,
                "account_new": account_status == "new",
                "description": description[:300],
                "competition_id": competition_id,
                "competition_label": label,
                "amount": round(amount, 2),
                "qty": qty,
                # el mensual solo aplica al primer año: es el mismo dato que el
                # anual dividido en 12, y proyectado a 2030 ya no significa nada
                "monthly_amount": monthly if line_year == base_year else None,
                "escalation_pct": escalation,
                "notes": notes,
                "series_id": series_id,
                "kind": kind,
                "status": "new",
                "existing_id": None,
                "message": None,
            }
            if amount < 0:
                warnings.append({"row": row_no, "message": f"Monto negativo ({amount}) en «{description}»"})

            match = by_id.get(line_id) if line_id else None
            if account_error:
                item["status"] = "error"
                item["message"] = account_error
            elif line_id and not match:
                item["status"] = "error"
                item["message"] = "La línea del ID no existe (o no la puedes ver)"
            elif match:
                item["status"], item["existing_id"] = "update", match["id"]
                item["series_id"] = match["series_id"]
            else:
                found = by_key.get(_row_key(line_year, row_department, account_code,
                                            competition_id, description))
                if found:
                    item["status"], item["existing_id"] = "update", found["id"]
                    item["series_id"] = found["series_id"]

            if editable is not None and row_department not in editable:
                item["status"] = "error"
                item["message"] = f"Sin permiso de edición sobre «{row_department}»"
            items.append(item)

    if len(items) > MAX_ROWS:
        errors.append(f"La planilla tiene más de {MAX_ROWS} líneas — divídela en varias.")
        items = items[:MAX_ROWS]

    if not items and not skipped_cells:
        if _has_formulas(content, parsed["sheet"]):
            raise ValueError(
                "El archivo tiene fórmulas sin valores calculados. Ábrelo en Excel, "
                "guárdalo de nuevo y vuelve a subirlo."
            )
        raise ValueError(f"No encontré filas con montos en la hoja '{parsed['sheet']}'.")

    if not create_accounts:
        for item in items:
            if item.get("account_new") and item["status"] != "error":
                item["status"] = "error"
                item["message"] = f"La cuenta {item['account_code']} no existe"

    # ── qué se borraría en modo reemplazo ───────────────────────────────────
    deletions = []
    target_departments = {item["department_code"] for item in items if item["status"] != "error"}
    # Sin una sola línea importable no hay con qué reemplazar: borrar igual
    # dejaría el año vacío por un archivo que no se pudo leer.
    if replace and target_departments:
        keep = {item["existing_id"] for item in items if item["existing_id"]}
        for line in existing:
            if line["id"] in keep:
                continue
            if line["department_code"] not in target_departments:
                continue
            if line["year"] not in years:
                continue
            if editable is not None and line["department_code"] not in editable:
                continue
            deletions.append(line)

    return {
        "sheet": parsed["sheet"],
        "sheets": parsed["sheets"],
        "shape": parsed["shape"],
        "headerless": parsed["headerless"],
        "header_row": parsed["header_row"],
        "years": years,
        "columns": [{"header": c["header"], "role": c["role"], "year": c["year"]} for c in columns],
        "competitions": [
            {"label": label, **{k: v for k, v in match.items() if k != "suggested_id"},
             "suggested_id": match.get("suggested_id")}
            for label, match in resolved.items()
        ],
        "new_accounts": sorted(new_accounts.values(), key=lambda a: -a["amount"]),
        "departments": sorted({item["department_code"] for item in items}),
        "items": items,
        "skipped_cells": skipped_cells,
        "deletions": deletions,
        "ignored_rows": ignored_rows,
        "warnings": warnings,
        "errors": errors,
    }


def _totals(plan: dict) -> dict:
    def total(status):
        return round(sum(i["amount"] for i in plan["items"] if i["status"] == status), 2)
    return {
        "new": total("new"),
        "update": total("update"),
        "error": total("error"),
        "import": round(total("new") + total("update"), 2),
        "skipped": round(sum(c["amount"] for c in plan["skipped_cells"]), 2),
        "delete": len(plan["deletions"]),
    }


def _counts(plan: dict) -> dict:
    counts = {"new": 0, "update": 0, "error": 0}
    for item in plan["items"]:
        counts[item["status"]] = counts.get(item["status"], 0) + 1
    return counts


# ─── Preview / commit ────────────────────────────────────────────────────────
_PREVIEW_ROWS = 300


def preview(content: bytes, **kwargs) -> dict:
    """Qué haría el import. No escribe nada."""
    plan = _plan(content, **kwargs)
    return {
        "sheet": plan["sheet"], "sheets": plan["sheets"], "shape": plan["shape"],
        "headerless": plan["headerless"], "header_row": plan["header_row"],
        "years": plan["years"], "columns": plan["columns"],
        "competitions": plan["competitions"], "new_accounts": plan["new_accounts"],
        "departments": plan["departments"], "skipped_cells": plan["skipped_cells"][:_PREVIEW_ROWS],
        "deletions": [
            {"id": d["id"], "year": d["year"], "department_code": d["department_code"],
             "account_code": d["account_code"], "description": d["description"]}
            for d in plan["deletions"][:_PREVIEW_ROWS]
        ],
        "ignored_rows": plan["ignored_rows"], "warnings": plan["warnings"][:100],
        "errors": plan["errors"],
        "counts": _counts(plan), "totals": _totals(plan),
        "rows": [
            {k: item[k] for k in ("row", "year", "department_code", "account_code", "account_new",
                                  "description", "competition_id", "competition_label", "amount",
                                  "status", "message")}
            for item in plan["items"][:_PREVIEW_ROWS]
        ],
        "truncated": len(plan["items"]) > _PREVIEW_ROWS,
    }


_LINE_FIELDS = ("year", "department_code", "account_code", "competition_id", "kind",
                "description", "qty", "monthly_amount", "amount", "escalation_pct",
                "notes", "series_id")


def commit(content: bytes, *, created_by: Optional[str] = None, **kwargs) -> dict:
    """Escribe el plan: crea las cuentas nuevas, inserta, actualiza y borra."""
    plan = _plan(content, **kwargs)
    importable = [i for i in plan["items"] if i["status"] in ("new", "update")]

    # 1. cuentas nuevas, como provisorias — el plan de cuentas oficial de
    #    Finance todavía no llegó (BUDGET_MODULE.md §12)
    accounts_created = []
    wanted = {}
    for item in importable:
        if item["account_new"]:
            wanted.setdefault(item["account_code"], item["description"][:120])
    for code, label in wanted.items():
        if supabase.table("accounts").select("code").eq("code", code).execute().data:
            continue
        supabase.table("accounts").insert({
            "code": code, "label": label, "kind": kwargs.get("kind", "expense"),
            "pending_mapping": True, "active": True,
        }).execute()
        accounts_created.append(code)

    # 2. líneas
    created, updated = 0, 0
    inserts = []
    for item in importable:
        record = {k: item[k] for k in _LINE_FIELDS if item.get(k) is not None}
        if item["status"] == "update":
            record.pop("series_id", None)
            supabase.table("budget_lines").update(record).eq("id", item["existing_id"]).execute()
            updated += 1
        else:
            record["source"] = "manual"
            if created_by:
                record["created_by"] = created_by
            inserts.append(record)

    if inserts:
        # PostgREST exige el mismo juego de claves en todo el batch
        keys = set().union(*(r.keys() for r in inserts))
        inserts = [{k: r.get(k) for k in keys} for r in inserts]
        for start in range(0, len(inserts), 100):
            supabase.table("budget_lines").insert(inserts[start:start + 100]).execute()
            created += len(inserts[start:start + 100])

    # 3. reemplazo: lo que estaba cargado y no vino en la planilla
    deleted = 0
    for line in plan["deletions"]:
        supabase.table("budget_lines").delete().eq("id", line["id"]).execute()
        deleted += 1

    return {
        "created": created, "updated": updated, "deleted": deleted,
        "accounts_created": accounts_created,
        "skipped": len(plan["items"]) - len(importable),
        "skipped_cells": len(plan["skipped_cells"]),
        "totals": _totals(plan),
        "departments": plan["departments"],
    }


# ─── Export (la plantilla del import) ────────────────────────────────────────
_HEADERS = {
    "es": ["ID", "Cuenta", "Nombre de cuenta", "Descripción", "Departamento",
           "Competencia", "Cantidad", "Mensual", "Escalación %", "Monto", "Notas"],
    "en": ["ID", "Account", "Account name", "Description", "Department",
           "Competition", "Qty", "Monthly", "Escalation %", "Amount", "Notes"],
}


def _safe(value):
    """Texto que ninguna planilla vaya a interpretar como fórmula."""
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD):
        return "'" + value
    return value


def export_lines_xlsx(lines: list, year: int, lang: str = "es") -> bytes:
    """Las líneas de un año en el formato que el import vuelve a leer.

    Es la plantilla y el round-trip a la vez: se exporta, se edita en Excel y se
    vuelve a subir. La columna ID hace que la reimportación actualice esas
    mismas filas en vez de duplicarlas — sin ella habría que confiar en que
    nadie tocó la descripción.
    """
    headers = _HEADERS.get(lang, _HEADERS["es"])
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {year}"

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF0C2340")   # navy del DS
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for line in lines:
        ws.append([
            line.get("id"),
            _safe(line.get("account_code")),
            _safe(line.get("account_label")),
            _safe(line.get("description")),
            _safe(line.get("department_label") or line.get("department_code")),
            _safe(line.get("competition_name")),
            line.get("qty"),
            line.get("monthly_amount"),
            line.get("escalation_pct"),
            line.get("amount"),
            _safe(line.get("notes")),
        ])

    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 46)
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
