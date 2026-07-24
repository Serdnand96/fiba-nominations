"""Game & Practice Schedule workbook generator (official FIBA layout).

Builds the combined games + practices schedule as an .xlsx replicating the
FIBA Americas "Game & Practice Schedule" template: one block per day with
practice slots (Estadio / Sede de Entrenamiento columns), a PARTIDOS section
listing that day's games, game-day labels relative to the first game day
("DÍA DE PARTIDO -2", "DÍA DE PARTIDO 1", "DÍA DE DESCANSO", phase suffixes)
and a comments column fed from slot notes.

Two layers, both pure (no DB access) so they can be tested standalone:
- build_schedule_header() / build_schedule_days(): rows from the
  `competitions`, `game_schedule` and `training_slots` tables -> intermediate
  representation (header dict + list of day blocks).
- build_schedule_workbook(): intermediate representation -> xlsx bytes
  (openpyxl). This is the only Excel-writing code path in the backend; the
  PDF exports (docx-based) are a separate pipeline, see
  `.claude/skills/pdf-templates`.
"""
import io
from collections import Counter
from datetime import datetime, time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Palette / typography from the official FIBA template ─────────────────────
FONT_TITLE = "FIBA Regular"     # falls back silently if not installed
FONT_BODY = "Univers Light"
FONT_HEADER = "Univers 55 Roman"

DARK = "FF333F4F"     # day-block header (FECHA / label / COMENTARIOS)
GRAY = "FFBFBFBF"     # sub-headers (Inicia/Termina/…) and GRUPO/SEDE boxes
AMBER = "FFFFBF00"    # PARTIDOS divider
GREEN = "FFA8CF8E"    # special events (rehearsals, ceremonies)
GAMELBL = "FFD9D9D9"  # round label cells (Cuartos de Final #1, Final, …)
LIGHT = "FFD9D9D9"    # team cells on the training-venue side (white -15%)
BLACK = "FF000000"
WHITE = "FFFFFFFF"

DATE_FMT = "[$-409]dd\\-mmm\\-yy;@"
TIME_FMT = "h:mm"

COL_WIDTHS = {
    "A": 6.8, "B": 4.6, "C": 11.8, "D": 3.4, "E": 12.6, "F": 1.2,
    "G": 15.2, "H": 3.4, "I": 8.4, "J": 9.4, "K": 8.4, "L": 32.8,
}

THIN = Side(style="thin", color="FF808080")
MEDIUM = Side(style="medium", color=BLACK)

# leading chars that spreadsheet apps may interpret as a formula — values
# starting with these are forced to plain text (CWE-1236: the exported file
# is shared with third parties, so team/venue strings must never execute)
_FORMULA_LEAD = ("=", "+", "-", "@")

# team_label keywords that mark a slot as an "event" (full-width green row)
EVENT_KEYWORDS = ("ensayo", "ceremonia", "prueba", "media day", "rehearsal", "ceremony")

# venue value (training_slots.venue) that maps to the training-site column;
# anything else goes to the stadium column
TRAINING_VENUE_LABEL = "cancha de entrenamiento"

LABELS = {
    "es": {
        "city": "Ciudad, País",
        "group": "GRUPO",
        "venue": "SEDE",
        "training_venue": "SEDE DE ENTRENAMIENTO",
        "date": "FECHA",
        "comments": "COMENTARIOS",
        "starts": "Inicia",
        "ends": "Termina",
        "stadium": "Estadio",
        "training_site": "Sede de Entrenamiento",
        "games": "PARTIDOS",
        "sheet": "Game & Practice Schedule",
        "game_day": "DÍA DE PARTIDO {n}",
        "rest_day": "DÍA DE DESCANSO",
        "suffix_qf": " - CUARTOS DE FINAL",
        "suffix_sf": " - SEMIFINALES",
        "suffix_final": " - FINALES",
        "suffix_class": " - CLASIFICACIÓN",
        "round_qf": "Cuartos de Final #{k}",
        "round_sf": "Semifinal #{k}",
        "round_class": "Clasificación",
        "round_finals": ["Final", "3er Lugar", "5to Lugar", "7mo Lugar"],
        "and": " y ",
        "weekdays": ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES",
                     "SÁBADO", "DOMINGO"],
    },
    "en": {
        "city": "City, Country",
        "group": "GROUP",
        "venue": "VENUE",
        "training_venue": "TRAINING VENUE",
        "date": "DATE",
        "comments": "COMMENTS",
        "starts": "Starts",
        "ends": "Ends",
        "stadium": "Stadium",
        "training_site": "Training Venue",
        "games": "GAMES",
        "sheet": "Game & Practice Schedule",
        "game_day": "GAME DAY {n}",
        "rest_day": "REST DAY",
        "suffix_qf": " - QUARTERFINALS",
        "suffix_sf": " - SEMIFINALS",
        "suffix_final": " - FINALS",
        "suffix_class": " - CLASSIFICATION",
        "round_qf": "Quarterfinal #{k}",
        "round_sf": "Semifinal #{k}",
        "round_class": "Classification",
        "round_finals": ["Final", "3rd Place", "5th Place", "7th Place"],
        "and": " and ",
        "weekdays": ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY",
                     "SATURDAY", "SUNDAY"],
    },
}


# ── Small helpers ────────────────────────────────────────────────────────────

def _norm_date(value) -> str | None:
    """Normalize a date value to YYYY-MM-DD (or None)."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()[:10] or None


def _norm_time(value) -> str | None:
    """Normalize a time value to HH:MM (or None if not parseable)."""
    if value in (None, ""):
        return None
    if isinstance(value, time):
        return value.strftime("%H:%M")
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h, m = int(parts[0]), int(parts[1])
            if 0 <= h <= 23 and 0 <= m <= 59:
                return f"{h:02d}:{m:02d}"
        except ValueError:
            pass
    return None


def _time_sort_key(value: str | None) -> str:
    return value if value else "99:99"


def _game_number_key(value) -> tuple:
    """Sortable key for game_number, which may be str, int or None."""
    try:
        return (0, int(value))
    except (TypeError, ValueError):
        return (1, 0)


def _game_key(g: dict):
    """Stable identity for a game row (DB id, falling back to object id)."""
    return g.get("id") or id(g)


def _phase_key(phase: str | None) -> str:
    """Map a free-text phase to one of: group, qf, sf, class, finals.

    Order matters: "semifinal" contains "final", so semi is checked first.
    """
    p = (phase or "").lower()
    if "semi" in p:
        return "sf"
    if "quarter" in p or "cuarto" in p or "qf" in p:
        return "qf"
    if "class" in p or "clasif" in p:
        return "class"
    if "final" in p:
        return "finals"
    return "group"


def _is_event_label(label: str | None) -> bool:
    text = (label or "").lower()
    return any(k in text for k in EVENT_KEYWORDS)


def _join_names(names: list, lang: str) -> str:
    """Join team names FIBA-style: "Colombia, Paraguay y Argentina"."""
    names = [n for n in names if n]
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return ", ".join(names[:-1]) + LABELS[lang]["and"] + names[-1]


def _mode(values: list) -> str | None:
    """Most common non-empty value, ties broken by first appearance."""
    cleaned = [str(v).strip() for v in values if v and str(v).strip()]
    if not cleaned:
        return None
    counts = Counter(cleaned)
    return max(cleaned, key=lambda v: (counts[v], -cleaned.index(v)))


# ── Intermediate representation builders ─────────────────────────────────────

def build_schedule_header(competition: dict, games: list, main_venue=None,
                          training_venue=None, lang: str = "es") -> dict:
    """Header info (title, city, groups, venues) from competition + games."""
    cities = []
    for g in games:
        city, country = (g.get("city") or "").strip(), (g.get("country") or "").strip()
        if city and country:
            cities.append(f"{city}, {country}")
        elif city or country:
            cities.append(city or country)
    city = _mode(cities) or (competition.get("location") or "")

    # groups: teams per group_label, first-appearance order (group phase only)
    groups: dict = {}
    for g in sorted(games, key=lambda g: (_norm_date(g.get("date")) or "",
                                          _time_sort_key(_norm_time(g.get("time"))))):
        if _phase_key(g.get("phase")) != "group":
            continue
        label = (g.get("group_label") or "").strip()
        if not label:
            continue
        teams = groups.setdefault(label, [])
        for team in (g.get("team_a"), g.get("team_b")):
            team = (team or "").strip()
            if team and team not in teams:
                teams.append(team)

    def group_title(label: str) -> str:
        up = label.upper()
        if up.startswith(("GRUPO", "GROUP")):
            return up
        return f"{LABELS[lang]['group']} {up}"

    group_labels = sorted(groups.keys())[:2]
    group_a = group_b = None
    if group_labels:
        group_a = {"title": group_title(group_labels[0]),
                   "teams": _join_names(groups[group_labels[0]], lang)}
    if len(group_labels) > 1:
        group_b = {"title": group_title(group_labels[1]),
                   "teams": _join_names(groups[group_labels[1]], lang)}

    return {
        "title": competition.get("name") or "",
        "city": city,
        "group_a": group_a,
        "group_b": group_b,
        "main_venue": (main_venue or "").strip() or _mode([g.get("venue") for g in games]) or "TBC",
        "training_venue": (training_venue or "").strip() or "TBC",
    }


def _round_labels(games: list, lang: str) -> dict:
    """Map game id -> round label (Cuartos de Final #1, Final, …)."""
    L = LABELS[lang]
    by_phase: dict = {}
    ordered = sorted(games, key=lambda g: (_norm_date(g.get("date")) or "",
                                           _time_sort_key(_norm_time(g.get("time"))),
                                           _game_number_key(g.get("game_number"))))
    for g in ordered:
        by_phase.setdefault(_phase_key(g.get("phase")), []).append(g)

    labels: dict = {}
    for k, tpl in (("qf", L["round_qf"]), ("sf", L["round_sf"])):
        for i, g in enumerate(by_phase.get(k, []), start=1):
            labels[_game_key(g)] = tpl.format(k=i)
    for g in by_phase.get("class", []):
        labels[_game_key(g)] = L["round_class"]
    finals = by_phase.get("finals", [])
    # last final = Final, then 3rd place, 5th place… walking backwards
    for i, g in enumerate(reversed(finals)):
        labels[_game_key(g)] = L["round_finals"][i] if i < len(L["round_finals"]) else L["round_class"]
    return labels


def build_schedule_days(games: list, slots: list, lang: str = "es") -> list:
    """Build the per-day blocks (label, practices, games, comments)."""
    L = LABELS[lang]
    round_by_game = _round_labels(games, lang)

    games_by_date: dict = {}
    for g in games:
        d = _norm_date(g.get("date"))
        if d:
            games_by_date.setdefault(d, []).append(g)

    slots_by_date: dict = {}
    for s in slots:
        d = _norm_date(s.get("date"))
        if d:
            slots_by_date.setdefault(d, []).append(s)

    game_dates = sorted(games_by_date.keys())
    all_dates = sorted(set(game_dates) | set(slots_by_date.keys()))
    first_game = datetime.strptime(game_dates[0], "%Y-%m-%d") if game_dates else None

    days = []
    for d in all_dates:
        day_games = sorted(games_by_date.get(d, []),
                           key=lambda g: (_time_sort_key(_norm_time(g.get("time"))),
                                          _game_number_key(g.get("game_number"))))
        day_slots = sorted(slots_by_date.get(d, []),
                           key=lambda s: (_time_sort_key(_norm_time(s.get("start_time"))),
                                          _time_sort_key(_norm_time(s.get("end_time")))))

        # day label
        try:
            date_obj = datetime.strptime(d, "%Y-%m-%d")
        except ValueError:
            date_obj = None
        if day_games:
            n = game_dates.index(d) + 1
            label = L["game_day"].format(n=n)
            phases = {_phase_key(g.get("phase")) for g in day_games}
            for key, suffix in (("finals", "suffix_final"), ("sf", "suffix_sf"),
                                ("qf", "suffix_qf"), ("class", "suffix_class")):
                if key in phases:
                    label += L[suffix]
                    break
        elif first_game and date_obj and date_obj < first_game:
            label = L["game_day"].format(n=-(first_game - date_obj).days)
        elif first_game:
            label = L["rest_day"]
        elif date_obj:
            label = f"{L['weekdays'][date_obj.weekday()]} {date_obj.strftime('%d/%m')}"
        else:
            label = ""

        # practices: pair stadium/training slots that share the same times
        practices = []
        by_times: dict = {}
        times_order = []
        for s in day_slots:
            key = (_norm_time(s.get("start_time")), _norm_time(s.get("end_time")))
            if key not in by_times:
                by_times[key] = {"stadium": [], "training": []}
                times_order.append(key)
            venue = (s.get("venue") or "").strip().lower()
            side = "training" if venue == TRAINING_VENUE_LABEL else "stadium"
            by_times[key][side].append(s)

        comments = []
        for key in times_order:
            start_t, end_t = key
            pair = by_times[key]
            for i in range(max(len(pair["stadium"]), len(pair["training"]))):
                st = pair["stadium"][i] if i < len(pair["stadium"]) else None
                tr = pair["training"][i] if i < len(pair["training"]) else None
                st_label = (st or {}).get("team_label") or None
                tr_label = (tr or {}).get("team_label") or None
                practices.append({
                    "start": start_t,
                    "end": end_t,
                    "stadium": st_label,
                    "training_venue": tr_label,
                    "is_event": _is_event_label(st_label) or _is_event_label(tr_label),
                })
                for slot in (st, tr):
                    note = (slot or {}).get("notes")
                    note = str(note).strip() if note else ""
                    if note and note not in comments:
                        comments.append(note)

        days.append({
            "date": d,
            "label": label,
            "comments": "\n".join(f"- {c}" for c in comments) if comments else None,
            "practices": practices,
            "games": [{
                "time": _norm_time(g.get("time")),
                "round_label": round_by_game.get(_game_key(g)),
                "home": (g.get("team_a") or "").strip() or None,
                "away": (g.get("team_b") or "").strip() or None,
            } for g in day_games],
        })
    return days


# ── Workbook rendering ───────────────────────────────────────────────────────

def _fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def _cell_time(value):
    """HH:MM string -> datetime.time so Excel formats it via TIME_FMT."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return value


def _cell_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return value


class _Painter:
    """Tiny helper to write value + styles in one call."""

    def __init__(self, ws):
        self.ws = ws

    def cell(self, row, col, value=None, *, font=None, fill=None, align=None,
             numfmt=None, border=None):
        c = self.ws.cell(row=row, column=col)
        if value is not None:
            c.value = value
            if isinstance(value, str) and value.startswith(_FORMULA_LEAD):
                c.data_type = "s"
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if numfmt:
            c.number_format = numfmt
        if border:
            c.border = border
        return c

    def span(self, row, col_from, col_to, value=None, **kw):
        """Merged horizontal range; style applied to every cell so borders close."""
        self.ws.merge_cells(start_row=row, start_column=col_from,
                            end_row=row, end_column=col_to)
        first = self.cell(row, col_from, value, **kw)
        for col in range(col_from + 1, col_to + 1):
            self.cell(row, col, font=kw.get("font"), fill=kw.get("fill"),
                      align=kw.get("align"), border=kw.get("border"))
        return first

    def strip(self, row, col_from, col_to, value_col, value=None, **kw):
        """Unmerged strip: fill/border across the range, value on the wide
        column — same approach as the official template (text overflows)."""
        for col in range(col_from, col_to + 1):
            self.cell(row, col, fill=kw.get("fill"), border=kw.get("border"))
        return self.cell(row, value_col, value, **kw)


def build_schedule_workbook(header: dict, days: list, lang: str = "es") -> bytes:
    """Render the intermediate representation to xlsx bytes."""
    L = LABELS[lang]

    wb = Workbook()
    ws = wb.active
    ws.title = L["sheet"]
    ws.sheet_view.showGridLines = False
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    p = _Painter(ws)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    top_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

    body = Font(name=FONT_BODY, size=7.5)
    body_b = Font(name=FONT_BODY, size=7.5, bold=True)
    header_f = Font(name=FONT_HEADER, size=8, bold=True)
    white_b = Font(name=FONT_BODY, size=7.5, bold=True, color=WHITE)
    grid = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    # title bar
    p.span(1, 1, 12, header.get("title") or "",
           font=Font(name=FONT_TITLE, size=16, bold=True, color=WHITE),
           fill=_fill(BLACK), align=center)
    ws.row_dimensions[1].height = 22

    p.cell(2, 1, L["city"], font=Font(name=FONT_BODY, size=7.5, bold=True),
           align=center)
    p.cell(2, 3, header.get("city") or "", font=Font(name=FONT_BODY, size=9.5),
           align=left)

    # groups + venues boxes
    group_a, group_b = header.get("group_a"), header.get("group_b")
    p.span(4, 3, 9, (group_a or {}).get("title") or f"{L['group']} A",
           font=header_f, fill=_fill(GRAY), align=center, border=grid)
    p.span(5, 3, 9, (group_a or {}).get("teams") or "", font=header_f,
           align=left, border=grid)
    p.span(7, 3, 9, (group_b or {}).get("title") or f"{L['group']} B",
           font=header_f, fill=_fill(GRAY), align=center, border=grid)
    p.span(8, 3, 9, (group_b or {}).get("teams") or "", font=header_f,
           align=left, border=grid)
    p.span(4, 11, 12, L["venue"], font=header_f, fill=_fill(GRAY),
           align=center, border=grid)
    p.span(5, 11, 12, header.get("main_venue") or "", font=header_f,
           align=center, border=grid)
    p.span(7, 11, 12, L["training_venue"], font=header_f, fill=_fill(GRAY),
           align=center, border=grid)
    p.span(8, 11, 12, header.get("training_venue") or "", font=header_f,
           align=center, border=grid)

    # day blocks (contiguous, like the official template)
    row = 10
    for day in days:
        # header row: FECHA | day label | COMENTARIOS
        p.cell(row, 1, L["date"], font=white_b, fill=_fill(DARK),
               align=center, border=grid)
        p.cell(row, 2, font=white_b, fill=_fill(DARK), border=grid)
        p.span(row, 3, 11, day.get("label") or "", font=white_b,
               fill=_fill(DARK), align=center, border=grid)
        p.cell(row, 12, L["comments"], font=white_b, fill=_fill(DARK),
               align=center, border=grid)
        row += 1

        # sub-header: Inicia | Termina | Estadio | Sede de Entrenamiento
        sub = row
        p.cell(row, 3, L["starts"], font=body_b, fill=_fill(GRAY),
               align=center, border=grid)
        p.cell(row, 4, fill=_fill(GRAY), border=grid)
        p.cell(row, 5, L["ends"], font=body_b, fill=_fill(GRAY),
               align=center, border=grid)
        p.span(row, 6, 8, L["stadium"], font=body_b, fill=_fill(GRAY),
               align=center, border=grid)
        p.span(row, 9, 11, L["training_site"], font=body_b, fill=_fill(GRAY),
               align=center, border=grid)
        row += 1

        for pr in day.get("practices", []):
            stadium, training = pr.get("stadium"), pr.get("training_venue")
            is_event = bool(pr.get("is_event"))
            fill = _fill(GREEN) if is_event else None
            p.cell(row, 1, border=grid)
            p.cell(row, 2, border=grid)
            p.cell(row, 3, _cell_time(pr.get("start")), font=body,
                   align=center, numfmt=TIME_FMT, border=grid, fill=fill)
            p.cell(row, 4, border=grid, fill=fill)
            p.cell(row, 5, _cell_time(pr.get("end")), font=body,
                   align=center, numfmt=TIME_FMT, border=grid, fill=fill)
            if is_event:
                # rehearsal/ceremony: single green band across both venues
                p.span(row, 6, 11, stadium or training, font=body,
                       align=center, border=grid, fill=fill)
            else:
                # empty venue side rendered black ("blocked", like the FIBA
                # template); training-side team cells get the light gray tint
                p.strip(row, 6, 8, 7, stadium, font=body, align=center,
                        border=grid, fill=None if stadium else _fill(BLACK))
                p.strip(row, 9, 11, 10, training, font=body, align=center,
                        border=grid,
                        fill=_fill(LIGHT) if training else _fill(BLACK))
            row += 1

        games = day.get("games", [])
        if games:
            p.cell(row, 1, border=grid)
            p.cell(row, 2, border=grid)
            p.span(row, 3, 11, L["games"], font=body_b, fill=_fill(AMBER),
                   align=center, border=grid)
            p.cell(row, 12, border=grid)
            row += 1
            for g in games:
                p.cell(row, 1, border=grid)
                p.cell(row, 2, border=grid)
                p.cell(row, 3, _cell_time(g.get("time")), font=body_b,
                       align=center, numfmt=TIME_FMT, border=grid)
                lbl = g.get("round_label")
                p.span(row, 4, 5, lbl or None, font=body_b,
                       fill=_fill(GAMELBL) if lbl else None,
                       align=center, border=grid)
                p.strip(row, 6, 8, 7, g.get("home"), font=body,
                        align=center, border=grid)
                p.strip(row, 9, 11, 10, g.get("away"), font=body,
                        align=center, border=grid)
                row += 1

        # date on col A:B merged vertically over the whole block (same
        # construct as the official file's GAMES sheet)
        if row - 1 >= sub:
            ws.merge_cells(start_row=sub, start_column=1,
                           end_row=row - 1, end_column=2)
        p.cell(sub, 1, _cell_date(day.get("date")),
               font=Font(name=FONT_BODY, size=9, bold=True),
               align=center, numfmt=DATE_FMT)

        # comments column, merged down the block
        if row - 1 > sub:
            ws.merge_cells(start_row=sub, start_column=12,
                           end_row=row - 1, end_column=12)
        p.cell(sub, 12, day.get("comments") or None, font=body, align=top_wrap)
        for r in range(sub, row):
            p.cell(r, 12, border=grid)

        # medium bottom border closes the block
        for col in range(1, 13):
            c = ws.cell(row=row - 1, column=col)
            b = c.border
            c.border = Border(top=b.top, left=b.left, right=b.right,
                              bottom=MEDIUM)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
